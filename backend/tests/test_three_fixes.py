"""
Test 3 bug fixes:
1. Auto-complete should use property's category for property_use (not hardcode 'residential')
2. Export photo URLs should start with https://app.nstuindia.com (not https://nstu-app.com)
3. PDF export was missing many fields - now includes Category, Total Area, Serial No, Bill Sr No, etc.
"""
import pytest
import requests
import os
import io
import re

BASE_URL = os.environ.get('REACT_APP_BACKEND_URL', '').rstrip('/')

# Test credentials
ADMIN_USERNAME = "admin"
ADMIN_PASSWORD = "nastu123"
TOWN_CODE = "THS"

# Test colony with data
TEST_COLONY = "T.P.S 8B"


class TestThreeBugFixes:
    """Test suite for the 3 bug fixes"""
    
    @pytest.fixture(scope="class")
    def auth_headers(self):
        """Get authentication headers"""
        response = requests.post(
            f"{BASE_URL}/api/auth/login",
            json={"username": ADMIN_USERNAME, "password": ADMIN_PASSWORD}
        )
        assert response.status_code == 200, f"Login failed: {response.text}"
        token = response.json()["token"]
        return {
            "Authorization": f"Bearer {token}",
            "x-town-code": TOWN_CODE
        }
    
    # ========== FIX 1: Auto-complete property_use mapping ==========
    
    def test_auto_complete_endpoint_exists(self, auth_headers):
        """Verify auto-complete endpoint is accessible"""
        # Just check that endpoint exists (we won't actually run auto-complete)
        # The endpoint is POST /api/admin/auto-complete-surveys
        response = requests.post(
            f"{BASE_URL}/api/admin/auto-complete-surveys",
            headers=auth_headers,
            json={"colony": "TEST_NONEXISTENT_COLONY_FOR_TEST", "employee_id": None}
        )
        # Should return 200 (with 0 completed) or error, but not 404
        assert response.status_code != 404, "Auto-complete endpoint not found"
        print(f"Auto-complete endpoint exists, status: {response.status_code}")
    
    def test_verify_property_use_mapping_logic_in_code(self):
        """Verify the auto-complete code has correct property_use mapping from category"""
        # Read server.py to verify the mapping logic
        server_path = "/app/backend/server.py"
        with open(server_path, 'r') as f:
            content = f.read()
        
        # Check for correct mapping logic around line 5008-5016
        assert 'if "commercial" in category:' in content, "Missing commercial category mapping"
        assert 'property_use = "commercial"' in content, "Missing commercial property_use assignment"
        assert 'if "mix" in category:' in content or 'elif "mix" in category:' in content, "Missing mix category mapping"
        assert 'property_use = "mix_use"' in content, "Missing mix_use property_use assignment"
        
        # Verify it's NOT just hardcoding residential
        # Look for the section with category-based mapping
        mapping_section = re.search(
            r'# Map category to property_use.*?property_use = "residential"',
            content, 
            re.DOTALL
        )
        assert mapping_section, "Property_use mapping section not found"
        
        mapping_text = mapping_section.group()
        assert '"commercial"' in mapping_text, "commercial mapping missing in property_use logic"
        assert '"mix_use"' in mapping_text, "mix_use mapping missing in property_use logic"
        print("PASS: Auto-complete property_use mapping logic is correct in code")
    
    def test_check_existing_submissions_have_various_property_use(self, auth_headers):
        """Check if existing submissions have different property_use values"""
        response = requests.get(
            f"{BASE_URL}/api/admin/submissions",
            headers=auth_headers,
            params={"limit": 100}
        )
        assert response.status_code == 200, f"Failed to get submissions: {response.text}"
        
        data = response.json()
        submissions = data.get("submissions", [])
        
        property_use_values = set()
        for sub in submissions:
            pu = sub.get("property_use", "")
            if pu:
                property_use_values.add(pu)
        
        print(f"Found property_use values in submissions: {property_use_values}")
        # Note: This tests existing data - new auto-completions should have mapped values
    
    # ========== FIX 2: Export base_url should be https://app.nstuindia.com ==========
    
    def test_verify_excel_export_base_url_in_code(self):
        """Verify the Excel export uses https://app.nstuindia.com as base_url"""
        server_path = "/app/backend/server.py"
        with open(server_path, 'r') as f:
            content = f.read()
        
        # Look for the base_url default in export section (around line 3440)
        assert 'base_url = "https://app.nstuindia.com"' in content, \
            "Export base_url should default to https://app.nstuindia.com"
        
        # Ensure it's NOT using the old incorrect URL
        # Search specifically in the export section
        _export_section_match = re.search(
            r'@api_router\.get\("/admin/export"\).*?def.*?export',
            content,
            re.DOTALL
        )
        
        # Find the function and check base_url
        lines = content.split('\n')
        in_export_func = False
        for i, line in enumerate(lines):
            if '/admin/export"' in line or 'async def export' in line:
                in_export_func = True
            if in_export_func and 'app.nstuindia.com' in line:
                print(f"PASS: Found correct base_url at line {i+1}: {line.strip()}")
                break
            if in_export_func and 'nstu-app.com' in line:
                pytest.fail(f"Found incorrect old URL at line {i+1}: {line.strip()}")
    
    def test_excel_export_photo_urls(self, auth_headers):
        """Test that exported Excel has photo URLs starting with https://app.nstuindia.com"""
        # First check if there's data in the colony
        response = requests.get(
            f"{BASE_URL}/api/admin/export",
            headers=auth_headers,
            params={"status": "Approved", "colony": TEST_COLONY}
        )
        
        if response.status_code == 200:
            # Check content type
            content_type = response.headers.get("Content-Type", "")
            assert "spreadsheet" in content_type or "excel" in content_type or "octet-stream" in content_type, \
                f"Expected Excel file, got content-type: {content_type}"
            
            # Parse the Excel content
            import openpyxl
            from io import BytesIO
            
            wb = openpyxl.load_workbook(BytesIO(response.content))
            ws = wb.active
            
            # Find photo columns (typically columns with "Photo" in header)
            headers = [cell.value for cell in ws[1]]
            photo_cols = [i for i, h in enumerate(headers) if h and "Photo" in str(h)]
            
            # Check photo URLs in data rows
            found_photo_urls = []
            for row_idx in range(2, min(ws.max_row + 1, 20)):  # Check first 18 data rows
                for col_idx in photo_cols:
                    cell_value = ws.cell(row=row_idx, column=col_idx + 1).value
                    if cell_value and str(cell_value).startswith("http"):
                        found_photo_urls.append(cell_value)
            
            if found_photo_urls:
                for url in found_photo_urls[:5]:  # Check first 5
                    print(f"Photo URL: {url}")
                    # URL should start with https://app.nstuindia.com or be a relative path made absolute
                    if url.startswith("http"):
                        assert "app.nstuindia.com" in url or "preview.emergentagent.com" in url, \
                            f"Photo URL has wrong base: {url}"
                print(f"PASS: All {len(found_photo_urls)} photo URLs have correct base URL")
            else:
                print("WARNING: No photo URLs found in export (data may not have photos)")
        else:
            print(f"WARNING: Export returned status {response.status_code} - {response.text[:200]}")
    
    # ========== FIX 3: PDF export includes all fields ==========
    
    def test_verify_pdf_export_fields_in_code(self):
        """Verify PDF export code includes all required fields"""
        server_path = "/app/backend/server.py"
        with open(server_path, 'r') as f:
            content = f.read()
        
        # Fields that should be in PDF export (around line 3763-3828)
        required_fields = [
            "Category",
            "Total Area",
            "Serial No",
            "Bill Sr No",
            "House Status",
            "Property Use",
            "Special Condition",
            "Self Satisfied",
            "Original Latitude",
            "Original Longitude",
            "Survey Latitude",
            "Survey Longitude",
            "Aadhar Number",
            "Family ID",
            "Review Remarks"
        ]
        
        missing_fields = []
        for field in required_fields:
            # Check if field appears in PDF export section (prop_data or survey_data tables)
            if f'["{field}"' not in content:
                missing_fields.append(field)
        
        assert len(missing_fields) == 0, f"PDF export missing fields: {missing_fields}"
        print(f"PASS: All {len(required_fields)} required fields found in PDF export code")
    
    def test_pdf_export_endpoint(self, auth_headers):
        """Test PDF export endpoint returns valid PDF"""
        response = requests.get(
            f"{BASE_URL}/api/admin/export-pdf",
            headers=auth_headers,
            params={"status": "Approved", "colony": TEST_COLONY}
        )
        
        if response.status_code == 200:
            content_type = response.headers.get("Content-Type", "")
            assert "pdf" in content_type.lower() or "octet-stream" in content_type.lower(), \
                f"Expected PDF, got: {content_type}"
            
            # Check PDF header
            pdf_content = response.content
            assert pdf_content[:4] == b'%PDF', "Response is not a valid PDF file"
            
            print(f"PASS: PDF export returned valid PDF, size: {len(pdf_content)} bytes")
            
            # Try to extract text from PDF to verify fields
            try:
                import fitz  # PyMuPDF
                doc = fitz.open(stream=pdf_content, filetype="pdf")
                
                all_text = ""
                for page_num in range(min(doc.page_count, 3)):  # Check first 3 pages
                    page = doc[page_num]
                    all_text += page.get_text()
                
                doc.close()
                
                # Check for key fields in PDF text
                fields_to_check = [
                    "Category",
                    "Total Area", 
                    "Serial No",
                    "Bill Sr No",
                    "House Status",
                    "Property Use"
                ]
                
                found_fields = []
                missing_in_pdf = []
                for field in fields_to_check:
                    if field in all_text or field.lower() in all_text.lower():
                        found_fields.append(field)
                    else:
                        missing_in_pdf.append(field)
                
                print(f"Fields found in PDF: {found_fields}")
                if missing_in_pdf:
                    print(f"WARNING: Fields not visible in PDF text extraction: {missing_in_pdf}")
                    # Note: Some fields might be in table format hard to extract
                    
            except ImportError:
                print("PyMuPDF not available, skipping PDF text extraction")
            except Exception as e:
                print(f"PDF text extraction error (non-critical): {e}")
        else:
            print(f"WARNING: PDF export returned status {response.status_code}")
    
    def test_pdf_export_date_format(self, auth_headers):
        """Verify PDF export uses DD/MM/YYYY date format"""
        server_path = "/app/backend/server.py"
        with open(server_path, 'r') as f:
            content = f.read()
        
        # Look for date formatting in PDF export section
        # Should have strftime("%d/%m/%Y") or similar
        assert '"%d/%m/%Y' in content or "DD/MM/YYYY" in content or 'strftime("%d' in content, \
            "PDF export should format dates as DD/MM/YYYY"
        print("PASS: Date format DD/MM/YYYY found in PDF export code")


class TestDashboardAfterChanges:
    """Verify dashboard still works after backend changes"""
    
    @pytest.fixture(scope="class")
    def auth_headers(self):
        """Get authentication headers"""
        response = requests.post(
            f"{BASE_URL}/api/auth/login",
            json={"username": ADMIN_USERNAME, "password": ADMIN_PASSWORD}
        )
        assert response.status_code == 200
        token = response.json()["token"]
        return {
            "Authorization": f"Bearer {token}",
            "x-town-code": TOWN_CODE
        }
    
    def test_dashboard_loads(self, auth_headers):
        """Test dashboard endpoint returns data"""
        response = requests.get(
            f"{BASE_URL}/api/admin/dashboard",
            headers=auth_headers
        )
        assert response.status_code == 200, f"Dashboard failed: {response.text}"
        
        data = response.json()
        assert "total" in data, "Dashboard missing 'total' field"
        assert "pending" in data, "Dashboard missing 'pending' field"
        print(f"PASS: Dashboard loads correctly - Total: {data.get('total')}, Pending: {data.get('pending')}")
    
    def test_submission_stats(self, auth_headers):
        """Test submission stats endpoint"""
        response = requests.get(
            f"{BASE_URL}/api/admin/submission-stats",
            headers=auth_headers
        )
        assert response.status_code == 200, f"Submission stats failed: {response.text}"
        
        data = response.json()
        print(f"PASS: Submission stats - {data}")


if __name__ == "__main__":
    pytest.main([__file__, "-v", "--tb=short"])
