from fastapi import FastAPI, APIRouter, HTTPException, Depends, UploadFile, File, Form, Query, Header, Request
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, StreamingResponse, Response, JSONResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from starlette.middleware.gzip import GZipMiddleware
from motor.motor_asyncio import AsyncIOMotorClient, AsyncIOMotorGridFSBucket
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timezone, timedelta
import bcrypt
import jwt
from bson import ObjectId
import aiofiles
import csv
import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import json
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch, mm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image as RLImage, PageBreak
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.pdfgen import canvas
from PIL import Image as PILImage, ImageDraw, ImageFont
import tempfile
import fitz  # PyMuPDF for PDF processing
import re
import math
import base64
import asyncio
from functools import lru_cache
import time

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# ============== SIMPLE IN-MEMORY CACHE ==============
class SimpleCache:
    def __init__(self, ttl_seconds=60):
        self.cache: Dict[str, Any] = {}
        self.timestamps: Dict[str, float] = {}
        self.ttl = ttl_seconds
    
    def get(self, key: str):
        if key in self.cache:
            if time.time() - self.timestamps[key] < self.ttl:
                return self.cache[key]
            else:
                del self.cache[key]
                del self.timestamps[key]
        return None
    
    def set(self, key: str, value: Any):
        self.cache[key] = value
        self.timestamps[key] = time.time()
    
    def clear(self):
        self.cache.clear()
        self.timestamps.clear()

# Cache instances (60 second TTL for map data)
map_cache = SimpleCache(ttl_seconds=60)
colonies_cache = SimpleCache(ttl_seconds=300)  # 5 minutes for colonies list

# MongoDB connection with optimized settings for performance
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(
    mongo_url,
    maxPoolSize=50,  # Increase connection pool
    minPoolSize=10,
    maxIdleTimeMS=30000,
    serverSelectionTimeoutMS=5000
)
db = client[os.environ['DB_NAME']]

# GridFS for file storage in database
fs_bucket = AsyncIOMotorGridFSBucket(db)

# ============== DATABASE INDEXES FOR PERFORMANCE ==============
async def create_indexes():
    """Create MongoDB indexes for faster queries"""
    try:
        # Properties collection indexes
        await db.properties.create_index("id", unique=True, background=True)
        await db.properties.create_index("batch_id", background=True)
        await db.properties.create_index("ward", background=True)
        await db.properties.create_index("colony", background=True)
        await db.properties.create_index("status", background=True)
        await db.properties.create_index("assigned_employee_id", background=True)
        await db.properties.create_index("assigned_employee_ids", background=True)  # NEW: For array field
        await db.properties.create_index([("latitude", 1), ("longitude", 1)], background=True)
        await db.properties.create_index("serial_number", background=True)
        await db.properties.create_index("bill_sr_no", background=True)
        await db.properties.create_index("property_id", background=True)  # NEW: For property lookup
        # Compound indexes for common query patterns
        await db.properties.create_index([("ward", 1), ("status", 1)], background=True)
        await db.properties.create_index([("assigned_employee_id", 1), ("status", 1)], background=True)
        await db.properties.create_index([("assigned_employee_id", 1), ("status", 1), ("serial_number", 1)], background=True)  # NEW: Optimized for surveyor map
        
        # Users collection indexes
        await db.users.create_index("id", unique=True, background=True)
        await db.users.create_index("username", unique=True, background=True)
        await db.users.create_index("role", background=True)
        
        # Submissions collection indexes
        await db.submissions.create_index("id", unique=True, background=True)
        await db.submissions.create_index("property_record_id", background=True)
        await db.submissions.create_index("employee_id", background=True)
        await db.submissions.create_index("status", background=True)
        await db.submissions.create_index("submitted_at", background=True)
        await db.submissions.create_index([("employee_id", 1), ("submitted_at", -1)], background=True)
        
        # Bills collection indexes
        await db.bills.create_index("id", unique=True, background=True)
        await db.bills.create_index("colony", background=True)
        await db.bills.create_index("bill_sr_no", background=True)
        
        # Attendance collection indexes
        await db.attendance.create_index("employee_id", background=True)
        await db.attendance.create_index("date", background=True)
        await db.attendance.create_index([("employee_id", 1), ("date", 1)], unique=True, background=True)
        
        # Batches collection
        await db.batches.create_index("id", unique=True, background=True)
        await db.batches.create_index("status", background=True)
        
        # Generated PDFs collection - for storing generated PDF records
        await db.generated_pdfs.create_index("id", unique=True, background=True)
        await db.generated_pdfs.create_index("colony", background=True)
        await db.generated_pdfs.create_index("created_at", background=True)
        await db.generated_pdfs.create_index("pdf_type", background=True)
        
        logging.info("MongoDB indexes created successfully")
    except Exception as e:
        logging.warning(f"Index creation warning (may already exist): {e}")

# JWT Configuration
JWT_SECRET = os.environ.get('JWT_SECRET', 'nstu-property-tax-secret-key-2025')
JWT_ALGORITHM = "HS256"

# Create uploads directory (for backward compatibility and temp files)
UPLOAD_DIR = ROOT_DIR / "uploads"
UPLOAD_DIR.mkdir(exist_ok=True)

# Create the main app with increased body size limit for file uploads
app = FastAPI(title="NSTU Property Tax Manager")

# Add middleware to allow large file uploads (50MB max)
from starlette.middleware.base import BaseHTTPMiddleware
from starlette.requests import Request as StarletteRequest

# Add GZip compression for faster responses
app.add_middleware(GZipMiddleware, minimum_size=1000)

# Startup event to create indexes
@app.on_event("startup")
async def startup_event():
    await create_indexes()
    logging.info("Application started with optimized database indexes")

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# ============== GRIDFS HELPER FUNCTIONS ==============

async def save_file_to_gridfs(file_content: bytes, filename: str, content_type: str = "application/octet-stream") -> str:
    """Save file to MongoDB GridFS and return file_id"""
    file_id = await fs_bucket.upload_from_stream(
        filename,
        file_content,
        metadata={"content_type": content_type, "uploaded_at": datetime.now(timezone.utc).isoformat()}
    )
    return str(file_id)

async def get_file_from_gridfs(file_id: str) -> tuple:
    """Get file from GridFS by file_id, returns (content, filename, content_type)"""
    try:
        grid_out = await fs_bucket.open_download_stream(ObjectId(file_id))
        content = await grid_out.read()
        filename = grid_out.filename
        content_type = grid_out.metadata.get("content_type", "application/octet-stream") if grid_out.metadata else "application/octet-stream"
        return content, filename, content_type
    except Exception:
        return None, None, None

async def delete_file_from_gridfs(file_id: str) -> bool:
    """Delete file from GridFS"""
    try:
        await fs_bucket.delete(ObjectId(file_id))
        return True
    except Exception:
        return False

# ============== FILE SERVE ENDPOINTS ==============

@api_router.get("/file/{file_id}")
async def serve_file(file_id: str):
    """Serve file from GridFS by file_id"""
    try:
        content, filename, content_type = await get_file_from_gridfs(file_id)
        if content is None:
            raise HTTPException(status_code=404, detail="File not found")
        
        return Response(
            content=content,
            media_type=content_type,
            headers={
                "Content-Disposition": f'inline; filename="{filename}"',
                "Cache-Control": "public, max-age=86400"  # Cache for 24 hours
            }
        )
    except Exception as e:
        raise HTTPException(status_code=404, detail=f"File not found: {str(e)}")

# Legacy endpoint for backward compatibility with old /api/uploads/ URLs
@api_router.get("/uploads/{filename}")
async def serve_legacy_upload(filename: str):
    """Serve legacy files from UPLOAD_DIR for backward compatibility"""
    file_path = UPLOAD_DIR / filename
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="File not found")
    
    # Determine content type
    suffix = file_path.suffix.lower()
    content_types = {
        '.jpg': 'image/jpeg', '.jpeg': 'image/jpeg', '.png': 'image/png',
        '.gif': 'image/gif', '.pdf': 'application/pdf', '.xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    }
    content_type = content_types.get(suffix, 'application/octet-stream')
    
    return FileResponse(
        path=str(file_path),
        media_type=content_type,
        headers={"Cache-Control": "public, max-age=86400"}
    )

# ============== MODELS ==============

# Role options: ADMIN, SURVEYOR, SUPERVISOR, MC_OFFICER
# Permission options for SUPERVISOR and MC_OFFICER
AVAILABLE_PERMISSIONS = [
    "dashboard",      # View dashboard
    "bills",          # View/manage bills
    "properties",     # View properties
    "map",            # View property map
    "submissions",    # View submissions
    "approve",        # Approve/reject submissions
    "employees",      # View employees
    "attendance",     # View attendance
    "export",         # Export data
    "upload"          # Upload data
]

class UserCreate(BaseModel):
    username: str
    password: str
    name: str
    role: str = "SURVEYOR"  # ADMIN, SURVEYOR, SUPERVISOR, MC_OFFICER
    assigned_area: Optional[str] = None
    authority: Optional[str] = None  # For SUPERVISOR and MC_OFFICER roles
    permissions: Optional[List[str]] = None  # For SUPERVISOR and MC_OFFICER roles

class UserLogin(BaseModel):
    username: str
    password: str

class UserResponse(BaseModel):
    id: str
    username: str
    name: str
    role: str
    assigned_area: Optional[str] = None
    authority: Optional[str] = None
    permissions: Optional[List[str]] = None
    created_at: str

class TokenResponse(BaseModel):
    token: str
    user: UserResponse

class DatasetBatchCreate(BaseModel):
    name: str

class DatasetBatchResponse(BaseModel):
    id: str
    name: str
    uploaded_by: str
    uploaded_at: str
    status: str
    total_records: int

class PropertyResponse(BaseModel):
    id: str
    batch_id: str
    property_id: str
    owner_name: str
    mobile: str
    address: str
    total_area: Optional[str] = None
    amount: Optional[str] = None
    ward: Optional[str] = None
    assigned_employee_id: Optional[str] = None
    assigned_employee_name: Optional[str] = None
    status: str
    created_at: str

class AssignmentRequest(BaseModel):
    property_ids: List[str]
    employee_id: Optional[str] = None  # Single employee (backward compat)
    employee_ids: Optional[List[str]] = None  # Multiple employees (work together)

class BulkAssignmentRequest(BaseModel):
    area: str
    employee_id: Optional[str] = None  # Single employee (backward compat)
    employee_ids: Optional[List[str]] = None  # Multiple employees (work together)
    custom_distribution: Optional[Dict[str, int]] = None  # {employee_id: count} for custom distribution
    serial_from: Optional[int] = None  # Range assignment: start serial number
    serial_to: Optional[int] = None    # Range assignment: end serial number

class BulkUnassignRequest(BaseModel):
    area: str
    employee_id: Optional[str] = None  # If provided, only unassign this employee from area

class SubmissionApproval(BaseModel):
    submission_id: str
    action: str  # APPROVE or REJECT
    remarks: Optional[str] = None

class DashboardStats(BaseModel):
    total_properties: int
    completed: int
    pending: int
    in_progress: int
    rejected: int
    employees: int
    batches: int
    today_completed: int
    today_wards: int

class EmployeeProgress(BaseModel):
    employee_id: str
    employee_name: str
    role: str
    total_assigned: int
    completed: int
    pending: int
    today_completed: int
    overall_completed: int

# ============== HELPER FUNCTIONS ==============

def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode(), bcrypt.gensalt()).decode()

def verify_password(password: str, hashed: str) -> bool:
    return bcrypt.checkpw(password.encode(), hashed.encode())

def create_token(user_id: str, role: str) -> str:
    payload = {
        "user_id": user_id,
        "role": role,
        "exp": datetime.now(timezone.utc).timestamp() + 86400 * 7  # 7 days
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

async def get_current_user(authorization: str = Header(None)):
    if not authorization:
        raise HTTPException(status_code=401, detail="Token required")
    try:
        token = authorization
        if token.startswith("Bearer "):
            token = token[7:]
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        user = await db.users.find_one({"id": payload["user_id"]}, {"_id": 0})
        if not user:
            raise HTTPException(status_code=401, detail="User not found")
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")

def get_today_start():
    """Get the start of today in UTC"""
    now = datetime.now(timezone.utc)
    return now.replace(hour=0, minute=0, second=0, microsecond=0)

# ============== AUTH ROUTES ==============

@api_router.post("/auth/login", response_model=TokenResponse)
async def login(data: UserLogin):
    user = await db.users.find_one({"username": data.username}, {"_id": 0})
    if not user or not verify_password(data.password, user["password_hash"]):
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    token = create_token(user["id"], user["role"])
    return {
        "token": token,
        "user": {
            "id": user["id"],
            "username": user["username"],
            "name": user["name"],
            "role": user["role"],
            "assigned_area": user.get("assigned_area"),
            "authority": user.get("authority"),
            "permissions": user.get("permissions"),
            "created_at": user["created_at"]
        }
    }

@api_router.get("/auth/me")
async def get_me(current_user: dict = Depends(get_current_user)):
    # Get user's custom permissions from DB
    user_permissions = current_user.get("permissions") or []
    role = current_user["role"]
    
    # For Admin, grant all permissions
    if role == "ADMIN":
        user_permissions = AVAILABLE_PERMISSIONS.copy()
    
    # Build permissions object - check both role-based and custom permissions
    permissions = {
        "can_upload": role in UPLOAD_ROLES or "upload" in user_permissions,
        "can_export": role in EXPORT_ROLES or "export" in user_permissions,
        "can_edit_submissions": role in SUBMISSION_EDIT_ROLES or "approve" in user_permissions,
        "can_manage_users": role == "ADMIN",
        "can_download_performance": role in PERFORMANCE_DOWNLOAD_ROLES or "export" in user_permissions,
        "can_view_employees": role in ["ADMIN", "MC_OFFICER"] or "employees" in user_permissions,
        "can_view_attendance": role in ["ADMIN", "MC_OFFICER"] or "attendance" in user_permissions,
        "can_approve_reject": role == "ADMIN" or "approve" in user_permissions,
        # New granular permissions
        "can_view_dashboard": role == "ADMIN" or "dashboard" in user_permissions,
        "can_view_bills": role == "ADMIN" or "bills" in user_permissions,
        "can_view_properties": role == "ADMIN" or "properties" in user_permissions,
        "can_view_map": role == "ADMIN" or "map" in user_permissions,
        "can_view_submissions": role == "ADMIN" or "submissions" in user_permissions,
    }
    
    return {
        "id": current_user["id"],
        "username": current_user["username"],
        "name": current_user["name"],
        "role": current_user["role"],
        "assigned_area": current_user.get("assigned_area"),
        "authority": current_user.get("authority"),
        "permissions": permissions,
        "raw_permissions": user_permissions,
        "created_at": current_user["created_at"]
    }

# ============== FAST MAP ENDPOINTS (OPTIMIZED FOR 20+ CONCURRENT USERS) ==============

@api_router.get("/map/colonies")
async def get_colonies_list(current_user: dict = Depends(get_current_user)):
    """Fast endpoint to get list of colonies - CACHED"""
    cache_key = "colonies_list"
    cached = colonies_cache.get(cache_key)
    if cached:
        return cached
    
    # Get unique colonies from properties
    pipeline = [
        {"$match": {"latitude": {"$ne": None}, "longitude": {"$ne": None}}},
        {"$group": {"_id": {"$ifNull": ["$colony", "$ward"]}, "count": {"$sum": 1}}},
        {"$match": {"_id": {"$ne": None}}},
        {"$sort": {"_id": 1}}
    ]
    
    result = await db.properties.aggregate(pipeline).to_list(None)
    colonies = [{"name": r["_id"], "count": r["count"]} for r in result if r["_id"]]
    total = sum(c["count"] for c in colonies)
    
    response = {"colonies": colonies, "total": total}
    colonies_cache.set(cache_key, response)
    return response

@api_router.get("/map/properties")
async def get_map_properties(
    colony: Optional[str] = None,
    status: Optional[str] = None,
    hide_completed: bool = False,
    limit: int = 5000,  # Increased from 500 to handle larger colonies
    current_user: dict = Depends(get_current_user)
):
    """Fast lightweight endpoint for map markers - NO DUPLICATES, shows submission status"""
    
    query = {"latitude": {"$ne": None}, "longitude": {"$ne": None}}
    
    if colony and colony.strip():
        query["$or"] = [
            {"colony": {"$regex": f"^{colony}$", "$options": "i"}},
            {"ward": {"$regex": f"^{colony}$", "$options": "i"}}
        ]
    
    # Filter by status if provided
    if status and status.strip():
        query["status"] = status
    
    # Hide completed/approved if requested
    if hide_completed:
        if "status" not in query:
            query["status"] = {"$nin": ["Completed", "Approved", "In Progress"]}
    
    # For non-admin users, filter by assigned properties
    if current_user["role"] not in ["ADMIN", "SUPERVISOR", "MC_OFFICER"]:
        if "$or" in query:
            query["$and"] = [
                {"$or": query.pop("$or")},
                {"$or": [
                    {"assigned_employee_id": current_user["id"]},
                    {"assigned_employee_ids": current_user["id"]}
                ]}
            ]
        else:
            query["$or"] = [
                {"assigned_employee_id": current_user["id"]},
                {"assigned_employee_ids": current_user["id"]}
            ]
    
    # Ultra-minimal projection for maximum speed
    projection = {
        "_id": 0,
        "id": 1,
        "latitude": 1,
        "longitude": 1,
        "status": 1,
        "serial_number": 1,
        "bill_sr_no": 1,
        "property_id": 1,
        "owner_name": 1,
        "colony": 1,
        "ward": 1,
        "mobile": 1,
        "assigned_employee_id": 1,
        "assigned_employee_name": 1,
        "assigned_employee_ids": 1,
        "category": 1,
        "total_area": 1,
        "amount": 1,
        "address": 1
    }
    
    properties = await db.properties.find(query, projection).limit(limit).to_list(limit)
    
    # Remove duplicates - keep unique by property_id ONLY
    seen_property_ids = set()
    unique_properties = []
    
    for prop in properties:
        prop_id = prop.get("property_id", "")
        
        # Skip if duplicate property_id
        if prop_id and prop_id in seen_property_ids:
            continue
        
        if prop_id:
            seen_property_ids.add(prop_id)
        
        unique_properties.append(prop)
    
    return {
        "properties": unique_properties,
        "count": len(unique_properties),
        "total_before_dedup": len(properties)
    }

@api_router.get("/map/employee-properties")
async def get_employee_map_properties(
    hide_completed: bool = False,
    current_user: dict = Depends(get_current_user)
):
    """Fast lightweight endpoint for surveyor map - Optimized for speed"""
    
    query = {
        "$or": [
            {"assigned_employee_id": current_user["id"]},
            {"assigned_employee_ids": current_user["id"]}
        ],
        "latitude": {"$ne": None},
        "longitude": {"$ne": None}
    }
    
    # Hide completed if requested
    if hide_completed:
        query["status"] = {"$nin": ["Completed", "Approved"]}
    
    # OPTIMIZED: Minimal projection for fast loading - only essential fields
    projection = {
        "_id": 0,
        "id": 1,
        "latitude": 1,
        "longitude": 1,
        "status": 1,
        "serial_number": 1,
        "bill_sr_no": 1,
        "property_id": 1,
        "owner_name": 1,
        "colony": 1,
        "ward": 1,
        "mobile": 1,
        "amount": 1,
        "total_area": 1,
        "category": 1,
        "address": 1,
        "self_certified": 1
    }
    
    # OPTIMIZED: Use index hint and batch size for faster queries
    properties = await db.properties.find(
        query, 
        projection,
        batch_size=1000  # Faster batch processing
    ).sort([
        ("status", 1),
        ("serial_number", 1)
    ]).to_list(None)
    
    # OPTIMIZED: Faster deduplication using dict
    seen = {}
    unique_properties = []
    
    for prop in properties:
        prop_id = prop.get("property_id", "")
        if prop_id:
            if prop_id not in seen:
                seen[prop_id] = True
                unique_properties.append(prop)
        else:
            unique_properties.append(prop)
    
    return {
        "properties": unique_properties,
        "count": len(unique_properties)
    }

# Clear cache when properties are modified
async def clear_map_cache():
    """Call this when properties are added/modified"""
    map_cache.clear()
    colonies_cache.clear()

# Get submission by property ID
@api_router.get("/submission/by-property/{property_id}")
async def get_submission_by_property(
    property_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Get submission data for a specific property"""
    if current_user["role"] not in ["ADMIN", "SUPERVISOR", "MC_OFFICER"]:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Search by property_record_id or property_id
    submission = await db.submissions.find_one(
        {"$or": [
            {"property_record_id": property_id},
            {"property_id": property_id}
        ]},
        {"_id": 0}
    )
    
    if not submission:
        return {"submission": None}
    
    return {"submission": submission}

# ============== ADMIN USER ROUTES ==============

@api_router.post("/admin/users", response_model=UserResponse)
async def create_user(data: UserCreate, current_user: dict = Depends(get_current_user)):
    if current_user["role"] != "ADMIN":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    existing = await db.users.find_one({"username": data.username})
    if existing:
        raise HTTPException(status_code=400, detail="Username already exists")
    
    # Set permissions based on role
    user_permissions = None
    if data.role in ["SUPERVISOR", "MC_OFFICER"]:
        # Use provided permissions or default to basic view permissions
        if data.permissions:
            # Validate permissions
            user_permissions = [p for p in data.permissions if p in AVAILABLE_PERMISSIONS]
        else:
            # Default permissions for SUPERVISOR/MC_OFFICER
            user_permissions = ["dashboard", "properties", "map", "submissions"]
    elif data.role == "ADMIN":
        # Admin has all permissions
        user_permissions = AVAILABLE_PERMISSIONS.copy()
    
    user_doc = {
        "id": str(uuid.uuid4()),
        "username": data.username,
        "password_hash": hash_password(data.password),
        "name": data.name,
        "role": data.role,
        "assigned_area": data.assigned_area,
        "authority": data.authority if data.role in ["SUPERVISOR", "MC_OFFICER"] else None,
        "permissions": user_permissions,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.users.insert_one(user_doc)
    
    return {
        "id": user_doc["id"],
        "username": user_doc["username"],
        "name": user_doc["name"],
        "role": user_doc["role"],
        "assigned_area": user_doc["assigned_area"],
        "authority": user_doc["authority"],
        "permissions": user_doc["permissions"],
        "created_at": user_doc["created_at"]
    }

@api_router.get("/admin/users", response_model=List[UserResponse])
async def list_users(current_user: dict = Depends(get_current_user)):
    if current_user["role"] != "ADMIN":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    users = await db.users.find({}, {"_id": 0, "password_hash": 0}).to_list(1000)
    return users

@api_router.delete("/admin/users/{user_id}")
async def delete_user(user_id: str, current_user: dict = Depends(get_current_user)):
    if current_user["role"] != "ADMIN":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    result = await db.users.delete_one({"id": user_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    return {"message": "User deleted"}

class UpdateUserRequest(BaseModel):
    name: Optional[str] = None
    authority: Optional[str] = None
    permissions: Optional[List[str]] = None
    assigned_area: Optional[str] = None

@api_router.put("/admin/users/{user_id}")
async def update_user(user_id: str, data: UpdateUserRequest, current_user: dict = Depends(get_current_user)):
    """Update user details (Admin only)"""
    if current_user["role"] != "ADMIN":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Check if user exists
    user = await db.users.find_one({"id": user_id})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Build update dict
    update_data = {}
    if data.name:
        update_data["name"] = data.name
    if data.authority is not None:
        update_data["authority"] = data.authority
    if data.permissions is not None:
        update_data["permissions"] = data.permissions
    if data.assigned_area is not None:
        update_data["assigned_area"] = data.assigned_area
    
    if update_data:
        await db.users.update_one({"id": user_id}, {"$set": update_data})
    
    # Return updated user
    updated_user = await db.users.find_one({"id": user_id}, {"_id": 0, "password_hash": 0})
    return updated_user

class ResetPasswordRequest(BaseModel):
    new_password: str

@api_router.post("/admin/users/{user_id}/reset-password")
async def reset_user_password(user_id: str, data: ResetPasswordRequest, current_user: dict = Depends(get_current_user)):
    """Reset password for a user (Admin only)"""
    if current_user["role"] != "ADMIN":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    if len(data.new_password) < 6:
        raise HTTPException(status_code=400, detail="Password must be at least 6 characters")
    
    # Check if user exists
    user = await db.users.find_one({"id": user_id})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Hash the new password
    hashed_password = bcrypt.hashpw(data.new_password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
    
    # Update password_hash (the field used by login verification)
    await db.users.update_one(
        {"id": user_id},
        {"$set": {"password_hash": hashed_password}}
    )
    
    return {"message": f"Password reset successfully for {user['name']}"}

# ============== BATCH UPLOAD ROUTES ==============

@api_router.post("/admin/batch/upload")
async def upload_batch(
    file: UploadFile = File(...),
    batch_name: str = Form(...),
    authorization: str = Form(...)
):
    current_user = await get_current_user(authorization)
    if current_user["role"] not in ADMIN_ROLES:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Read file content
    content = await file.read()
    filename = file.filename.lower()
    
    properties = []
    
    # Check file type and parse accordingly
    if filename.endswith('.xlsx') or filename.endswith('.xls'):
        # Parse Excel file using openpyxl
        import openpyxl
        from io import BytesIO
        
        workbook = openpyxl.load_workbook(BytesIO(content), data_only=True)
        sheet = workbook.active
        
        # Get headers from first row
        headers = []
        for cell in sheet[1]:
            headers.append(str(cell.value).strip() if cell.value else "")
        
        # Create header mapping (case-insensitive)
        header_map = {h.lower(): i for i, h in enumerate(headers)}
        
        # Parse data rows starting from row 2
        serial_num = 1
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not any(row):  # Skip empty rows
                continue
            
            # Get values using header mapping
            def get_val(keys):
                for k in keys:
                    idx = header_map.get(k.lower())
                    if idx is not None and idx < len(row) and row[idx]:
                        return str(row[idx]).strip()
                return ""
            
            prop = {
                "id": str(uuid.uuid4()),
                "serial_number": serial_num,
                "property_id": get_val(["Property Id", "property_id", "PropertyID"]) or str(uuid.uuid4())[:8].upper(),
                "old_property_id": get_val(["Old Property Id", "old_property_id", "OldPropertyId"]),
                "owner_name": get_val(["Owner Name", "owner_name", "OwnerName"]) or "Unknown",
                "mobile": get_val(["Mobile", "mobile", "Mobile No", "Phone"]),
                "address": get_val(["Plot Address", "Address", "address", "plot_address"]),
                "colony": get_val(["Colony", "colony", "Area", "area"]),
                "ward": get_val(["Colony", "Ward", "ward", "area"]),
                "latitude": None,
                "longitude": None,
                "total_area": get_val(["Total Area (SqYard)", "Total Area", "total_area", "Area"]),
                "category": get_val(["Category", "category"]),
                "amount": get_val(["Outstanding", "Total Outstanding", "Amount", "amount"]) or "0",
                "financial_year": get_val(["Financial Year", "financial_year"]) or "2025-2026",
                "assigned_employee_id": None,
                "assigned_employee_name": None,
                "status": "Pending",
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            
            # Parse latitude/longitude if available
            lat_str = get_val(["Latitude", "latitude", "Lat"])
            lng_str = get_val(["Longitude", "longitude", "Long", "Lng"])
            if lat_str:
                try:
                    prop["latitude"] = float(lat_str)
                except:
                    pass
            if lng_str:
                try:
                    prop["longitude"] = float(lng_str)
                except:
                    pass
            
            properties.append(prop)
            serial_num += 1
    else:
        # Parse CSV file
        content_str = content.decode('utf-8')
        reader = csv.DictReader(io.StringIO(content_str))
        
        serial_num = 1
        for row in reader:
            prop = {
                "id": str(uuid.uuid4()),
                "serial_number": serial_num,
                "property_id": row.get("property_id") or row.get("Property Id") or row.get("PropertyID") or str(uuid.uuid4())[:8].upper(),
                "old_property_id": row.get("old_property_id") or row.get("Old Property Id") or "",
                "owner_name": row.get("owner_name") or row.get("Owner Name") or row.get("OwnerName") or "Unknown",
                "mobile": row.get("mobile") or row.get("Mobile") or row.get("Mobile No") or "",
                "address": row.get("address") or row.get("Address") or row.get("Plot Address") or row.get("plot_address") or "",
                "colony": row.get("Colony") or row.get("colony") or row.get("Area") or "",
                "ward": row.get("ward") or row.get("Ward") or row.get("Colony") or row.get("area") or "",
                "latitude": None,
                "longitude": None,
                "total_area": row.get("total_area") or row.get("Total Area") or row.get("Total Area (SqYard)") or "",
                "category": row.get("Category") or row.get("category") or "",
                "amount": row.get("amount") or row.get("Amount") or row.get("Outstanding") or "0",
                "financial_year": row.get("Financial Year") or row.get("financial_year") or "2025-2026",
                "assigned_employee_id": None,
                "assigned_employee_name": None,
                "status": "Pending",
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            
            # Parse latitude/longitude
            lat_str = row.get("Latitude") or row.get("latitude")
            lng_str = row.get("Longitude") or row.get("longitude")
            if lat_str:
                try:
                    prop["latitude"] = float(lat_str)
                except:
                    pass
            if lng_str:
                try:
                    prop["longitude"] = float(lng_str)
                except:
                    pass
            
            properties.append(prop)
            serial_num += 1
    
    if not properties:
        raise HTTPException(status_code=400, detail="No valid properties found in file")
    
    # Create batch
    batch_doc = {
        "id": str(uuid.uuid4()),
        "name": batch_name,
        "uploaded_by": current_user["id"],
        "uploaded_at": datetime.now(timezone.utc).isoformat(),
        "status": "ACTIVE",
        "total_records": len(properties)
    }
    await db.batches.insert_one(batch_doc)
    
    # Add batch_id to properties and insert
    for prop in properties:
        prop["batch_id"] = batch_doc["id"]
    
    if properties:
        await db.properties.insert_many(properties)
    
    return {
        "batch_id": batch_doc["id"],
        "name": batch_doc["name"],
        "total_records": len(properties),
        "message": f"Successfully uploaded {len(properties)} properties"
    }

@api_router.get("/admin/batches", response_model=List[DatasetBatchResponse])
async def list_batches(current_user: dict = Depends(get_current_user)):
    if current_user["role"] != "ADMIN":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    batches = await db.batches.find({"status": {"$ne": "DELETED"}}, {"_id": 0}).to_list(100)
    return batches

@api_router.post("/admin/batch/{batch_id}/archive")
async def archive_batch(batch_id: str, current_user: dict = Depends(get_current_user)):
    if current_user["role"] != "ADMIN":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    result = await db.batches.update_one(
        {"id": batch_id},
        {"$set": {"status": "ARCHIVED"}}
    )
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Batch not found")
    return {"message": "Batch archived"}

@api_router.delete("/admin/batch/{batch_id}")
async def delete_batch(batch_id: str, current_user: dict = Depends(get_current_user)):
    if current_user["role"] != "ADMIN":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    await db.properties.delete_many({"batch_id": batch_id})
    await db.submissions.delete_many({"batch_id": batch_id})
    await db.batches.delete_one({"id": batch_id})
    
    return {"message": "Batch and all related data deleted"}

# ============== ROLE DEFINITIONS ==============
# Roles with admin-level access (can modify data)
ADMIN_ROLES = ["ADMIN", "SUPERVISOR"]
# Roles that can view admin dashboard (including MC_OFFICER with limited access)
ADMIN_VIEW_ROLES = ["ADMIN", "SUPERVISOR", "MC_OFFICER"]
# Roles that can export data (PDF/Excel)
EXPORT_ROLES = ["ADMIN", "MC_OFFICER"]
# Roles that can upload data
UPLOAD_ROLES = ["ADMIN", "SUPERVISOR"]
# Roles that can edit submissions
SUBMISSION_EDIT_ROLES = ["ADMIN"]
# Roles that can download employee performance
PERFORMANCE_DOWNLOAD_ROLES = ["ADMIN"]

# ============== PROPERTY ROUTES ==============

@api_router.get("/admin/properties")
async def list_properties(
    batch_id: Optional[str] = None,
    ward: Optional[str] = None,
    status: Optional[str] = None,
    employee_id: Optional[str] = None,
    search: Optional[str] = None,
    page: int = 1,
    limit: int = 50,
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] not in ADMIN_VIEW_ROLES:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    query = {}
    if batch_id and batch_id.strip():
        query["batch_id"] = batch_id
    if ward and ward.strip():
        query["ward"] = ward
    if status and status.strip():
        query["status"] = status
    if employee_id and employee_id.strip():
        query["assigned_employee_id"] = employee_id
    if search:
        query["$or"] = [
            {"property_id": {"$regex": search, "$options": "i"}},
            {"owner_name": {"$regex": search, "$options": "i"}},
            {"mobile": {"$regex": search, "$options": "i"}}
        ]
    
    # Optimized projection for faster queries
    projection = {
        "_id": 0,
        "id": 1,
        "property_id": 1,
        "owner_name": 1,
        "mobile": 1,
        "address": 1,
        "colony": 1,
        "ward": 1,
        "latitude": 1,
        "longitude": 1,
        "status": 1,
        "serial_number": 1,
        "bill_sr_no": 1,
        "amount": 1,
        "category": 1,
        "total_area": 1,
        "assigned_employee_id": 1,
        "assigned_employee_name": 1,
        "assigned_employee_ids": 1,
        "batch_id": 1,
        "created_at": 1
    }
    
    skip = (page - 1) * limit
    total = await db.properties.count_documents(query)
    properties = await db.properties.find(query, projection).sort("serial_number", 1).skip(skip).limit(limit).to_list(limit)
    
    return {
        "properties": properties,
        "total": total,
        "page": page,
        "pages": (total + limit - 1) // limit
    }

@api_router.post("/admin/assign")
async def assign_properties(data: AssignmentRequest, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ADMIN_ROLES:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Support both single employee_id and multiple employee_ids
    new_emp_ids = data.employee_ids if data.employee_ids else ([data.employee_id] if data.employee_id else [])
    
    if not new_emp_ids:
        raise HTTPException(status_code=400, detail="At least one employee must be selected")
    
    # Get all selected employees
    new_employees = await db.users.find({"id": {"$in": new_emp_ids}}, {"_id": 0}).to_list(None)
    if not new_employees:
        raise HTTPException(status_code=404, detail="No employees found")
    
    # Process each property to ADD new employees to existing assignments
    updated_count = 0
    for prop_id in data.property_ids:
        # Get existing property to check current assignments
        prop = await db.properties.find_one({"id": prop_id}, {"_id": 0})
        if not prop:
            continue
        
        # Get existing assigned employee IDs (or empty list)
        existing_emp_ids = prop.get("assigned_employee_ids") or []
        if prop.get("assigned_employee_id") and prop["assigned_employee_id"] not in existing_emp_ids:
            existing_emp_ids.append(prop["assigned_employee_id"])
        
        # Merge new employees with existing (avoid duplicates)
        combined_emp_ids = list(set(existing_emp_ids + new_emp_ids))
        
        # Get all employee names for the combined list
        all_employees = await db.users.find({"id": {"$in": combined_emp_ids}}, {"_id": 0, "id": 1, "name": 1}).to_list(None)
        combined_names = ", ".join([emp["name"] for emp in all_employees])
        
        # Update the property with merged assignments
        await db.properties.update_one(
            {"id": prop_id},
            {"$set": {
                "assigned_employee_ids": combined_emp_ids,
                "assigned_employee_id": combined_emp_ids[0] if combined_emp_ids else None,
                "assigned_employee_name": combined_names
            }}
        )
        updated_count += 1
    
    new_employee_names = ", ".join([emp["name"] for emp in new_employees])
    return {"message": f"Added {new_employee_names} to {updated_count} properties"}

@api_router.post("/admin/assign-bulk")
async def bulk_assign_by_ward(data: BulkAssignmentRequest, current_user: dict = Depends(get_current_user)):
    if current_user["role"] != "ADMIN":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Support both single employee_id and multiple employee_ids
    new_emp_ids = data.employee_ids if data.employee_ids else ([data.employee_id] if data.employee_id else [])
    
    if not new_emp_ids:
        raise HTTPException(status_code=400, detail="At least one employee must be selected")
    
    # Get all selected employees
    new_employees = await db.users.find({"id": {"$in": new_emp_ids}}, {"_id": 0}).to_list(None)
    if not new_employees:
        raise HTTPException(status_code=404, detail="No employees found")
    
    emp_name_map = {emp["id"]: emp["name"] for emp in new_employees}
    
    # Build query - just filter by area, serial filtering done in Python
    query = {"ward": data.area}
    
    # Get ALL properties in the ward/area (serial filtering done in Python to handle N-prefix)
    properties = await db.properties.find(query, {"_id": 0, "id": 1, "serial_number": 1, "bill_sr_no": 1, "assigned_employee_ids": 1, "assigned_employee_id": 1}).to_list(None)
    
    # RANGE ASSIGNMENT: Filter by serial number range (handles N-prefix)
    if data.serial_from is not None and data.serial_to is not None:
        filtered_props = []
        for prop in properties:
            serial = prop.get("serial_number") or 0
            bill_sr_str = str(prop.get("bill_sr_no") or "").strip()
            
            # Try to get numeric serial
            bill_sr_num = 0
            try:
                bill_sr_num = int(bill_sr_str)
            except:
                pass
            
            # Handle N-prefix serial numbers (e.g., N45, N584, N123)
            n_serial = 0
            if bill_sr_str.upper().startswith("N"):
                try:
                    n_serial = int(bill_sr_str[1:])  # Extract number after "N"
                except:
                    pass
            
            # Use the best available serial: actual serial > bill_sr_no as number > N-prefix number
            effective_serial = serial if serial > 0 else (bill_sr_num if bill_sr_num > 0 else n_serial)
            
            # Include if the effective serial is within range
            if effective_serial > 0 and data.serial_from <= effective_serial <= data.serial_to:
                filtered_props.append(prop)
        
        properties = filtered_props
    
    if not properties:
        raise HTTPException(status_code=404, detail=f"No properties found in {data.area}" + (f" with serial {data.serial_from}-{data.serial_to}" if data.serial_from else ""))
    
    # Check if custom distribution is provided
    if data.custom_distribution:
        # Custom distribution: assign specific count to each employee
        updated_count = 0
        prop_index = 0
        
        for emp_id, count in data.custom_distribution.items():
            if emp_id not in new_emp_ids:
                continue
                
            emp_name = emp_name_map.get(emp_id, "Unknown")
            
            # Assign 'count' properties to this employee
            for i in range(int(count)):
                if prop_index >= len(properties):
                    break
                    
                prop = properties[prop_index]
                
                # Get existing assigned employee IDs
                existing_emp_ids = prop.get("assigned_employee_ids") or []
                if prop.get("assigned_employee_id") and prop["assigned_employee_id"] not in existing_emp_ids:
                    existing_emp_ids.append(prop["assigned_employee_id"])
                
                # Add only this employee (not all)
                combined_emp_ids = list(set(existing_emp_ids + [emp_id]))
                
                # Get all employee names
                all_employees = await db.users.find({"id": {"$in": combined_emp_ids}}, {"_id": 0, "id": 1, "name": 1}).to_list(None)
                combined_names = ", ".join([e["name"] for e in all_employees])
                
                await db.properties.update_one(
                    {"id": prop["id"]},
                    {"$set": {
                        "assigned_employee_ids": combined_emp_ids,
                        "assigned_employee_id": combined_emp_ids[0] if combined_emp_ids else None,
                        "assigned_employee_name": combined_names
                    }}
                )
                updated_count += 1
                prop_index += 1
        
        # Build distribution summary
        dist_summary = ", ".join([f"{emp_name_map.get(eid, 'Unknown')}: {cnt}" for eid, cnt in data.custom_distribution.items()])
        return {"message": f"Assigned {updated_count} properties in {data.area} (Distribution: {dist_summary})"}
    
    else:
        # Default: assign ALL employees to ALL properties (existing behavior)
        updated_count = 0
        for prop in properties:
            existing_emp_ids = prop.get("assigned_employee_ids") or []
            if prop.get("assigned_employee_id") and prop["assigned_employee_id"] not in existing_emp_ids:
                existing_emp_ids.append(prop["assigned_employee_id"])
            
            combined_emp_ids = list(set(existing_emp_ids + new_emp_ids))
            
            all_employees = await db.users.find({"id": {"$in": combined_emp_ids}}, {"_id": 0, "id": 1, "name": 1}).to_list(None)
            combined_names = ", ".join([emp["name"] for emp in all_employees])
            
            await db.properties.update_one(
                {"id": prop["id"]},
                {"$set": {
                    "assigned_employee_ids": combined_emp_ids,
                    "assigned_employee_id": combined_emp_ids[0] if combined_emp_ids else None,
                    "assigned_employee_name": combined_names
                }}
            )
            updated_count += 1
        
        # Update all assigned employees with the area
        for emp_id in new_emp_ids:
            await db.users.update_one(
                {"id": emp_id},
                {"$set": {"assigned_area": data.area}}
            )
        
        new_employee_names = ", ".join([emp["name"] for emp in new_employees])
        range_info = f" (Serial {data.serial_from}-{data.serial_to})" if data.serial_from else ""
        return {"message": f"Assigned {new_employee_names} to {updated_count} properties in {data.area}{range_info}"}

@api_router.post("/admin/unassign-bulk")
async def bulk_unassign_by_ward(data: BulkUnassignRequest, current_user: dict = Depends(get_current_user)):
    """Bulk unassign properties by area"""
    if current_user["role"] not in ADMIN_ROLES:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    query = {"ward": data.area}
    
    # If specific employee, only unassign that employee
    if data.employee_id:
        employee = await db.users.find_one({"id": data.employee_id}, {"_id": 0, "name": 1})
        emp_name = employee["name"] if employee else "Unknown"
        
        # Find properties assigned to this employee in this area
        properties = await db.properties.find({
            "ward": data.area,
            "$or": [
                {"assigned_employee_id": data.employee_id},
                {"assigned_employee_ids": data.employee_id}
            ]
        }, {"_id": 0, "id": 1, "assigned_employee_ids": 1, "assigned_employee_id": 1}).to_list(None)
        
        updated_count = 0
        for prop in properties:
            existing_emp_ids = prop.get("assigned_employee_ids") or []
            
            # Remove this employee from the list
            new_emp_ids = [eid for eid in existing_emp_ids if eid != data.employee_id]
            
            if new_emp_ids:
                # Still has other employees
                all_employees = await db.users.find({"id": {"$in": new_emp_ids}}, {"_id": 0, "name": 1}).to_list(None)
                combined_names = ", ".join([e["name"] for e in all_employees])
                
                await db.properties.update_one(
                    {"id": prop["id"]},
                    {"$set": {
                        "assigned_employee_ids": new_emp_ids,
                        "assigned_employee_id": new_emp_ids[0],
                        "assigned_employee_name": combined_names
                    }}
                )
            else:
                # No employees left, clear all
                await db.properties.update_one(
                    {"id": prop["id"]},
                    {"$set": {
                        "assigned_employee_ids": [],
                        "assigned_employee_id": None,
                        "assigned_employee_name": None
                    }}
                )
            updated_count += 1
        
        return {"message": f"Removed {emp_name} from {updated_count} properties in {data.area}"}
    
    else:
        # Unassign ALL employees from ALL properties in this area
        result = await db.properties.update_many(
            query,
            {"$set": {
                "assigned_employee_ids": [],
                "assigned_employee_id": None,
                "assigned_employee_name": None
            }}
        )
        
        return {"message": f"Unassigned all employees from {result.modified_count} properties in {data.area}"}

# ============== UNASSIGN PROPERTIES ==============
class UnassignRequest(BaseModel):
    property_ids: List[str]
    employee_id: Optional[str] = None  # If provided, only unassign this employee. If not, unassign all.

@api_router.post("/admin/unassign")
async def unassign_properties(data: UnassignRequest, current_user: dict = Depends(get_current_user)):
    """Unassign employee(s) from properties"""
    if current_user["role"] not in ADMIN_ROLES:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    if not data.property_ids:
        raise HTTPException(status_code=400, detail="No properties selected")
    
    updated_count = 0
    
    for prop_id in data.property_ids:
        prop = await db.properties.find_one({"id": prop_id}, {"_id": 0})
        if not prop:
            continue
        
        if data.employee_id:
            # Unassign specific employee
            existing_emp_ids = prop.get("assigned_employee_ids") or []
            if prop.get("assigned_employee_id"):
                if prop["assigned_employee_id"] not in existing_emp_ids:
                    existing_emp_ids.append(prop["assigned_employee_id"])
            
            # Remove the specified employee
            new_emp_ids = [eid for eid in existing_emp_ids if eid != data.employee_id]
            
            if new_emp_ids:
                # Get remaining employee names
                remaining_employees = await db.users.find({"id": {"$in": new_emp_ids}}, {"_id": 0, "id": 1, "name": 1}).to_list(None)
                remaining_names = ", ".join([emp["name"] for emp in remaining_employees])
                
                await db.properties.update_one(
                    {"id": prop_id},
                    {"$set": {
                        "assigned_employee_ids": new_emp_ids,
                        "assigned_employee_id": new_emp_ids[0],
                        "assigned_employee_name": remaining_names
                    }}
                )
            else:
                # No employees left - clear all assignment fields
                await db.properties.update_one(
                    {"id": prop_id},
                    {"$set": {
                        "assigned_employee_ids": [],
                        "assigned_employee_id": None,
                        "assigned_employee_name": None
                    }}
                )
        else:
            # Unassign ALL employees
            await db.properties.update_one(
                {"id": prop_id},
                {"$set": {
                    "assigned_employee_ids": [],
                    "assigned_employee_id": None,
                    "assigned_employee_name": None
                }}
            )
        
        updated_count += 1
    
    if data.employee_id:
        emp = await db.users.find_one({"id": data.employee_id}, {"_id": 0, "name": 1})
        emp_name = emp["name"] if emp else "Employee"
        return {"message": f"Unassigned {emp_name} from {updated_count} properties"}
    else:
        return {"message": f"Unassigned all employees from {updated_count} properties"}

@api_router.post("/admin/unassign-by-employee")
async def unassign_all_properties_from_employee(
    employee_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Unassign ALL properties from a specific employee (when they leave)"""
    if current_user["role"] not in ADMIN_ROLES:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Get employee name
    employee = await db.users.find_one({"id": employee_id}, {"_id": 0, "name": 1})
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    # Find all properties assigned to this employee
    properties = await db.properties.find({
        "$or": [
            {"assigned_employee_id": employee_id},
            {"assigned_employee_ids": employee_id}
        ]
    }, {"_id": 0, "id": 1, "assigned_employee_ids": 1, "assigned_employee_id": 1}).to_list(None)
    
    updated_count = 0
    for prop in properties:
        existing_emp_ids = prop.get("assigned_employee_ids") or []
        if prop.get("assigned_employee_id") and prop["assigned_employee_id"] not in existing_emp_ids:
            existing_emp_ids.append(prop["assigned_employee_id"])
        
        # Remove the employee
        new_emp_ids = [eid for eid in existing_emp_ids if eid != employee_id]
        
        if new_emp_ids:
            remaining_employees = await db.users.find({"id": {"$in": new_emp_ids}}, {"_id": 0, "id": 1, "name": 1}).to_list(None)
            remaining_names = ", ".join([emp["name"] for emp in remaining_employees])
            
            await db.properties.update_one(
                {"id": prop["id"]},
                {"$set": {
                    "assigned_employee_ids": new_emp_ids,
                    "assigned_employee_id": new_emp_ids[0],
                    "assigned_employee_name": remaining_names
                }}
            )
        else:
            await db.properties.update_one(
                {"id": prop["id"]},
                {"$set": {
                    "assigned_employee_ids": [],
                    "assigned_employee_id": None,
                    "assigned_employee_name": None
                }}
            )
        updated_count += 1
    
    return {
        "message": f"Unassigned {employee['name']} from {updated_count} properties",
        "unassigned_count": updated_count
    }

class BulkDeleteRequest(BaseModel):
    property_ids: List[str]

@api_router.post("/admin/properties/bulk-delete")
async def bulk_delete_properties(data: BulkDeleteRequest, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ADMIN_ROLES:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    if not data.property_ids:
        raise HTTPException(status_code=400, detail="No properties selected for deletion")
    
    # Delete associated submissions first
    await db.submissions.delete_many({"property_record_id": {"$in": data.property_ids}})
    
    # Delete the properties
    result = await db.properties.delete_many({"id": {"$in": data.property_ids}})
    
    return {
        "message": f"Successfully deleted {result.deleted_count} properties",
        "deleted_count": result.deleted_count
    }

@api_router.post("/admin/properties/delete-all")
async def delete_all_properties(
    batch_id: Optional[str] = None,
    ward: Optional[str] = None,
    status: Optional[str] = None,
    employee_id: Optional[str] = None,
    search: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Delete all properties matching the given filters. If no filters, deletes ALL properties."""
    if current_user["role"] not in ADMIN_ROLES:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Build query based on filters
    query = {}
    if batch_id and batch_id.strip():
        query["batch_id"] = batch_id
    if ward and ward.strip():
        query["ward"] = ward
    if status and status.strip():
        query["status"] = status
    if employee_id and employee_id.strip():
        query["assigned_employee_id"] = employee_id
    if search and search.strip():
        query["$or"] = [
            {"property_id": {"$regex": search, "$options": "i"}},
            {"owner_name": {"$regex": search, "$options": "i"}},
            {"mobile": {"$regex": search, "$options": "i"}}
        ]
    
    # Get count first
    count = await db.properties.count_documents(query)
    
    if count == 0:
        return {"message": "No properties found to delete", "deleted_count": 0}
    
    # Get all property IDs to delete submissions
    properties = await db.properties.find(query, {"id": 1, "_id": 0}).to_list(None)
    property_ids = [p["id"] for p in properties]
    
    # Delete associated submissions first
    await db.submissions.delete_many({"property_record_id": {"$in": property_ids}})
    
    # Delete the properties
    result = await db.properties.delete_many(query)
    
    return {
        "message": f"Successfully deleted {result.deleted_count} properties",
        "deleted_count": result.deleted_count
    }

@api_router.post("/admin/properties/delete-colony")
async def delete_colony_properties(
    colony: str = Form(...),
    keep_surveyed: bool = Form(True),
    current_user: dict = Depends(get_current_user)
):
    """Delete all properties of a specific colony. Option to keep surveyed properties."""
    if current_user["role"] not in ADMIN_ROLES:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Find properties in this colony
    query = {"$or": [
        {"colony": {"$regex": f"^{colony}$", "$options": "i"}},
        {"ward": {"$regex": f"^{colony}$", "$options": "i"}}
    ]}
    
    properties = await db.properties.find(query, {"id": 1, "_id": 0}).to_list(None)
    property_ids = [p["id"] for p in properties]
    
    if not property_ids:
        return {"message": "No properties found in this colony", "deleted_count": 0}
    
    # If keep_surveyed, exclude properties with submissions
    ids_to_delete = property_ids
    if keep_surveyed:
        submissions = await db.submissions.find(
            {"property_record_id": {"$in": property_ids}},
            {"property_record_id": 1, "_id": 0}
        ).to_list(None)
        surveyed_ids = set(s["property_record_id"] for s in submissions)
        ids_to_delete = [pid for pid in property_ids if pid not in surveyed_ids]
    
    if not ids_to_delete:
        return {"message": "All properties in this colony have surveys. Nothing deleted.", "deleted_count": 0, "kept_surveyed": len(property_ids)}
    
    # Delete properties
    result = await db.properties.delete_many({"id": {"$in": ids_to_delete}})
    
    # Clear cache
    await clear_map_cache()
    
    return {
        "message": f"Deleted {result.deleted_count} properties from {colony}",
        "deleted_count": result.deleted_count,
        "kept_surveyed": len(property_ids) - len(ids_to_delete) if keep_surveyed else 0
    }

@api_router.post("/admin/properties/delete-duplicates")
async def delete_duplicate_properties(
    colony: str = Form(None),
    current_user: dict = Depends(get_current_user)
):
    """
    Delete duplicate properties but KEEP the ones with survey submissions.
    Duplicates are identified by: property_id OR (owner_name + mobile)
    """
    if current_user["role"] not in ADMIN_ROLES:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Build query
    query = {}
    if colony and colony.strip():
        query["$or"] = [
            {"colony": {"$regex": f"^{colony}$", "$options": "i"}},
            {"ward": {"$regex": f"^{colony}$", "$options": "i"}}
        ]
    
    # Get all properties
    properties = await db.properties.find(query, {"_id": 0}).to_list(None)
    
    if not properties:
        return {"message": "No properties found", "deleted_count": 0}
    
    # Get all submissions to know which properties have surveys
    all_submissions = await db.submissions.find({}, {"property_record_id": 1, "_id": 0}).to_list(None)
    surveyed_property_ids = set(s["property_record_id"] for s in all_submissions)
    
    # Group properties by property_id and by owner+mobile
    by_property_id = {}
    by_owner_mobile = {}
    
    for prop in properties:
        pid = prop.get("property_id", "")
        owner = (prop.get("owner_name") or "").strip().upper()
        mobile = (prop.get("mobile") or "").strip()
        prop_uuid = prop.get("id")
        has_survey = prop_uuid in surveyed_property_ids
        
        # Group by property_id
        if pid:
            if pid not in by_property_id:
                by_property_id[pid] = []
            by_property_id[pid].append({"uuid": prop_uuid, "has_survey": has_survey, "prop": prop})
        
        # Group by owner + mobile
        if owner and mobile:
            key = f"{owner}_{mobile}"
            if key not in by_owner_mobile:
                by_owner_mobile[key] = []
            by_owner_mobile[key].append({"uuid": prop_uuid, "has_survey": has_survey, "prop": prop})
    
    # Find duplicates to delete (keep ones with survey, delete others)
    ids_to_delete = set()
    
    # Check property_id duplicates
    for pid, items in by_property_id.items():
        if len(items) > 1:
            # Keep the one with survey, or the first one if none have survey
            has_survey_items = [i for i in items if i["has_survey"]]
            if has_survey_items:
                # Keep all with survey, delete others
                for item in items:
                    if not item["has_survey"]:
                        ids_to_delete.add(item["uuid"])
            else:
                # Keep only the first one
                for item in items[1:]:
                    ids_to_delete.add(item["uuid"])
    
    # Check owner+mobile duplicates
    for key, items in by_owner_mobile.items():
        if len(items) > 1:
            has_survey_items = [i for i in items if i["has_survey"]]
            if has_survey_items:
                for item in items:
                    if not item["has_survey"]:
                        ids_to_delete.add(item["uuid"])
            else:
                for item in items[1:]:
                    ids_to_delete.add(item["uuid"])
    
    if not ids_to_delete:
        return {"message": "No duplicate properties found", "deleted_count": 0}
    
    # Delete duplicates
    result = await db.properties.delete_many({"id": {"$in": list(ids_to_delete)}})
    
    # Clear cache
    await clear_map_cache()
    
    return {
        "message": f"Deleted {result.deleted_count} duplicate properties (kept surveyed ones)",
        "deleted_count": result.deleted_count,
        "total_properties": len(properties),
        "remaining": len(properties) - result.deleted_count
    }

@api_router.post("/admin/properties/arrange-by-route")
async def arrange_properties_by_route(
    ward: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Arrange properties by GPS route using nearest neighbor algorithm"""
    if current_user["role"] not in ADMIN_ROLES:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    query = {"latitude": {"$ne": None}, "longitude": {"$ne": None}}
    if ward and ward.strip():
        query["ward"] = ward
    
    # Get all properties with GPS
    properties = await db.properties.find(query, {"_id": 0}).to_list(None)
    
    if not properties:
        raise HTTPException(status_code=404, detail="No properties with GPS found")
    
    # Sort by GPS route using nearest neighbor algorithm
    sorted_props = []
    remaining = list(properties)
    
    if remaining:
        # Start from first property
        sorted_props.append(remaining.pop(0))
        
        while remaining:
            last = sorted_props[-1]
            last_lat, last_lon = last['latitude'], last['longitude']
            
            # Find nearest neighbor
            nearest_idx = 0
            nearest_dist = float('inf')
            
            for i, prop in enumerate(remaining):
                dist = haversine_distance(last_lat, last_lon, prop['latitude'], prop['longitude'])
                if dist < nearest_dist:
                    nearest_dist = dist
                    nearest_idx = i
            
            sorted_props.append(remaining.pop(nearest_idx))
    
    # Update serial numbers based on route order
    for i, prop in enumerate(sorted_props):
        await db.properties.update_one(
            {"id": prop["id"]},
            {"$set": {"serial_number": i + 1, "route_ordered": True}}
        )
    
    return {
        "message": f"Arranged {len(sorted_props)} properties by GPS route",
        "total_arranged": len(sorted_props)
    }

@api_router.post("/admin/properties/save-arranged")
async def save_arranged_data(
    ward: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Save the current arrangement as the permanent serial numbers"""
    if current_user["role"] not in ADMIN_ROLES:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    query = {}
    if ward and ward.strip():
        query["ward"] = ward
    
    # Get properties sorted by current serial_number
    properties = await db.properties.find(query, {"_id": 0}).sort("serial_number", 1).to_list(None)
    
    if not properties:
        raise HTTPException(status_code=404, detail="No properties found")
    
    # Re-assign serial numbers to ensure they're consecutive
    for i, prop in enumerate(properties):
        await db.properties.update_one(
            {"id": prop["id"]},
            {"$set": {"serial_number": i + 1, "arrangement_saved": True, "saved_at": datetime.now(timezone.utc).isoformat()}}
        )
    
    return {
        "message": f"Saved arrangement for {len(properties)} properties",
        "total_saved": len(properties)
    }

@api_router.post("/admin/properties/download-pdf")
async def download_properties_pdf(
    ward: Optional[str] = None,
    sn_position: str = "top-right",
    sn_font_size: int = 48,
    sn_color: str = "red",
    current_user: dict = Depends(get_current_user)
):
    """Generate PDF with property list arranged by serial number"""
    if current_user["role"] not in ADMIN_ROLES:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    query = {}
    if ward and ward.strip():
        query["ward"] = ward
    
    # Get properties sorted by serial_number
    properties = await db.properties.find(query, {"_id": 0}).sort("serial_number", 1).to_list(None)
    
    if not properties:
        raise HTTPException(status_code=404, detail="No properties found")
    
    # Generate PDF
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    colony_name = ward.replace(" ", "_") if ward else "all"
    pdf_filename = f"properties_{colony_name}_{timestamp}.pdf"
    pdf_path = UPLOAD_DIR / pdf_filename
    
    # Ensure uploads directory exists
    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    
    # Create PDF document
    doc = SimpleDocTemplate(str(pdf_path), pagesize=A4)
    styles = getSampleStyleSheet()
    elements = []
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        alignment=TA_CENTER,
        spaceAfter=20
    )
    elements.append(Paragraph(f"Property List - {ward or 'All Colonies'}", title_style))
    elements.append(Paragraph(f"Total: {len(properties)} properties | Generated: {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles['Normal']))
    elements.append(Spacer(1, 20))
    
    # Table data
    table_data = [['SN', 'Property ID', 'Owner Name', 'Mobile', 'Category', 'Area', 'Amount']]
    
    for prop in properties:
        table_data.append([
            str(prop.get('serial_number', '-')),
            prop.get('property_id', '-'),
            prop.get('owner_name', '-')[:20] if prop.get('owner_name') else '-',
            prop.get('mobile', '-'),
            prop.get('category', '-')[:10] if prop.get('category') else '-',
            prop.get('total_area', '-'),
            f"₹{prop.get('amount', '0')}"
        ])
    
    # Create table
    table = Table(table_data, colWidths=[30, 70, 100, 80, 60, 50, 60])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
    ]))
    
    elements.append(table)
    doc.build(elements)
    
    # Verify file was created
    if not pdf_path.exists():
        raise HTTPException(status_code=500, detail="PDF generation failed")
    
    return {
        "message": f"Generated PDF with {len(properties)} properties",
        "filename": pdf_filename,
        "download_url": f"/api/uploads/{pdf_filename}"
    }

# Direct PDF download endpoint - More reliable for VPS
@api_router.get("/admin/properties/download-pdf/{filename}")
async def get_pdf_file(filename: str, current_user: dict = Depends(get_current_user)):
    """Direct download of generated PDF file"""
    if current_user["role"] not in ADMIN_VIEW_ROLES:
        raise HTTPException(status_code=403, detail="Access denied")
    
    file_path = UPLOAD_DIR / filename
    if not file_path.exists():
        raise HTTPException(status_code=404, detail=f"File not found: {filename}")
    
    return FileResponse(
        path=str(file_path),
        media_type='application/pdf',
        filename=filename,
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"'
        }
    )

@api_router.get("/admin/wards")
async def list_wards(current_user: dict = Depends(get_current_user)):
    if current_user["role"] != "ADMIN":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    wards = await db.properties.distinct("ward")
    wards = [w for w in wards if w]
    return {"wards": wards}

# ============== DASHBOARD ROUTES ==============

@api_router.get("/admin/dashboard")
async def admin_dashboard(
    date: str = None,  # Optional date filter (YYYY-MM-DD format, empty = all time)
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] not in ADMIN_VIEW_ROLES:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Build date filter for submissions
    date_filter = {}
    if date:
        # Filter submissions by date
        date_start = f"{date}T00:00:00"
        date_end = f"{date}T23:59:59"
        date_filter = {"submitted_at": {"$gte": date_start, "$lte": date_end}}
    
    # Property counts (always all time for total)
    total = await db.properties.count_documents({})
    approved = await db.properties.count_documents({"status": "Approved"})
    completed = await db.properties.count_documents({"status": "Completed"})  # Surveyed but not yet approved
    pending = await db.properties.count_documents({"status": "Pending"})
    rejected = await db.properties.count_documents({"status": "Rejected"})
    employees = await db.users.count_documents({"role": {"$ne": "ADMIN"}})
    
    # Get unique colonies count
    colonies = await db.properties.distinct("colony")
    colonies_count = len([c for c in colonies if c])
    
    return {
        "total": total,
        "approved": approved + completed,  # Combined: Approved + Completed (both mean survey done)
        "pending": pending,
        "rejected": rejected,
        "employees": employees,
        "colonies": colonies_count
    }

@api_router.get("/admin/employee-progress")
async def get_employee_progress(
    date: str = None,  # Optional date filter
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] not in ADMIN_VIEW_ROLES:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    today_start = get_today_start().isoformat()
    
    employees = await db.users.find({"role": {"$ne": "ADMIN"}}, {"_id": 0}).to_list(100)
    progress = []
    
    for emp in employees:
        total = await db.properties.count_documents({"assigned_employee_id": emp["id"]})
        
        # Count approved properties (Completed + Approved = done)
        approved = await db.properties.count_documents({
            "assigned_employee_id": emp["id"],
            "status": {"$in": ["Completed", "Approved"]}
        })
        
        # Today's completed for this employee
        today_completed = await db.submissions.count_documents({
            "employee_id": emp["id"],
            "submitted_at": {"$gte": today_start},
            "status": {"$ne": "Rejected"}
        })
        
        # If date filter provided, use that instead
        if date:
            date_start = f"{date}T00:00:00"
            date_end = f"{date}T23:59:59"
            today_completed = await db.submissions.count_documents({
                "employee_id": emp["id"],
                "submitted_at": {"$gte": date_start, "$lte": date_end},
                "status": {"$ne": "Rejected"}
            })
        
        # Get assigned colonies for this employee
        assigned_colonies = await db.properties.distinct("ward", {"assigned_employee_id": emp["id"]})
        assigned_colonies = [c for c in assigned_colonies if c]  # Filter None
        
        progress.append({
            "employee_id": emp["id"],
            "employee_name": emp["name"],
            "role": emp["role"],
            "total_assigned": total,
            "completed": approved,  # Using approved (done) count
            "pending": total - approved,
            "today_completed": today_completed,
            "assigned_colonies": assigned_colonies,
            "colony_count": len(assigned_colonies)
        })
    
    return progress

@api_router.get("/admin/employee-progress/{employee_id}/colonies")
async def get_employee_colony_progress(employee_id: str, current_user: dict = Depends(get_current_user)):
    """Get detailed colony-wise progress for a specific employee"""
    if current_user["role"] not in ADMIN_VIEW_ROLES:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Get employee details
    employee = await db.users.find_one({"id": employee_id}, {"_id": 0})
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    # Get assigned colonies
    pipeline = [
        {"$match": {"assigned_employee_id": employee_id}},
        {"$group": {
            "_id": "$ward",
            "total": {"$sum": 1},
            "completed": {"$sum": {"$cond": [{"$in": ["$status", ["Completed", "Approved"]]}, 1, 0]}},
            "pending": {"$sum": {"$cond": [{"$eq": ["$status", "Pending"]}, 1, 0]}},
            "rejected": {"$sum": {"$cond": [{"$eq": ["$status", "Rejected"]}, 1, 0]}}
        }},
        {"$sort": {"_id": 1}}
    ]
    
    colony_stats = await db.properties.aggregate(pipeline).to_list(None)
    
    colonies = []
    for c in colony_stats:
        if c["_id"]:  # Skip None colony
            percentage = round((c["completed"] / c["total"]) * 100) if c["total"] > 0 else 0
            colonies.append({
                "colony": c["_id"],
                "total": c["total"],
                "completed": c["completed"],
                "pending": c["pending"],
                "rejected": c["rejected"],
                "percentage": percentage
            })
    
    return {
        "employee_id": employee_id,
        "employee_name": employee["name"],
        "role": employee["role"],
        "colonies": colonies,
        "total_colonies": len(colonies)
    }

@api_router.post("/admin/employee/remove-from-colony")
async def remove_employee_from_colony(
    employee_id: str = Form(...),
    colony: str = Form(...),
    current_user: dict = Depends(get_current_user)
):
    """Remove an employee from a specific colony - unassign all properties in that colony"""
    if current_user["role"] not in ADMIN_ROLES:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Check employee exists
    employee = await db.users.find_one({"id": employee_id}, {"_id": 0})
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    # Count properties to be unassigned
    count = await db.properties.count_documents({
        "assigned_employee_id": employee_id,
        "ward": colony
    })
    
    if count == 0:
        raise HTTPException(status_code=404, detail=f"No properties found for {employee['name']} in {colony}")
    
    # Unassign properties in this colony
    result = await db.properties.update_many(
        {
            "assigned_employee_id": employee_id,
            "ward": colony
        },
        {
            "$set": {
                "assigned_employee_id": None,
                "assigned_employee_name": None,
                "assignment_date": None
            }
        }
    )
    
    return {
        "message": f"Removed {employee['name']} from {colony}",
        "properties_unassigned": result.modified_count
    }

# ============== SUBMISSIONS ROUTES ==============

@api_router.get("/admin/areas")
async def list_areas(current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ADMIN_ROLES:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Get unique areas/wards from properties
    areas = await db.properties.distinct("ward")
    # Filter out None/empty values and sort
    areas = sorted([a for a in areas if a])
    
    return {"areas": areas}

@api_router.get("/admin/submission-stats")
async def get_submission_stats(
    date: str = None,  # Optional date filter (YYYY-MM-DD format, empty = all time)
    current_user: dict = Depends(get_current_user)
):
    """Get submission statistics for dashboard with optional date filter"""
    if current_user["role"] not in ADMIN_ROLES:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Build date filter
    query = {}
    if date:
        date_start = f"{date}T00:00:00"
        date_end = f"{date}T23:59:59"
        query["submitted_at"] = {"$gte": date_start, "$lte": date_end}
    
    total = await db.submissions.count_documents(query)
    
    pending_query = {**query, "status": "Pending"}
    approved_query = {**query, "status": "Approved"}
    completed_query = {**query, "status": "Completed"}
    rejected_query = {**query, "status": "Rejected"}
    
    pending = await db.submissions.count_documents(pending_query)
    approved = await db.submissions.count_documents(approved_query)
    completed = await db.submissions.count_documents(completed_query)
    rejected = await db.submissions.count_documents(rejected_query)
    
    return {
        "total": total,
        "pending": pending + completed,  # Pending review = Pending + Completed (not yet approved)
        "approved": approved,
        "rejected": rejected
    }

@api_router.get("/admin/submissions")
async def list_submissions(
    batch_id: Optional[str] = None,
    employee_id: Optional[str] = None,
    status: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    colony: Optional[str] = None,  # Colony filter
    search: Optional[str] = None,  # Search by serial number, property ID, owner name
    page: int = 1,
    limit: int = 50,
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] not in ADMIN_VIEW_ROLES:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    query = {}
    if batch_id and batch_id.strip():
        query["batch_id"] = batch_id
    if employee_id and employee_id.strip():
        query["employee_id"] = employee_id
    if status and status.strip():
        query["status"] = status
    if date_from:
        query["submitted_at"] = {"$gte": date_from}
    if date_to:
        if "submitted_at" in query:
            query["submitted_at"]["$lte"] = date_to
        else:
            query["submitted_at"] = {"$lte": date_to}
    
    # If search is provided, first find matching property IDs
    search_property_ids = None
    if search and search.strip():
        search_term = search.strip()
        # Search in properties collection
        search_query = {
            "$or": [
                {"id": {"$regex": search_term, "$options": "i"}},
                {"owner_name": {"$regex": search_term, "$options": "i"}},
                {"mobile": {"$regex": search_term, "$options": "i"}},
                {"bill_sr_no": {"$regex": search_term, "$options": "i"}}
            ]
        }
        # Try to search by serial number if it's a number
        try:
            serial_num = int(search_term)
            search_query["$or"].append({"serial_number": serial_num})
        except ValueError:
            pass
        
        matching_properties = await db.properties.find(search_query, {"id": 1, "_id": 0}).to_list(None)
        search_property_ids = [p["id"] for p in matching_properties]
        
        if search_property_ids:
            if "property_record_id" in query:
                # Intersect with existing filter
                query["property_record_id"]["$in"] = list(set(query["property_record_id"]["$in"]) & set(search_property_ids))
            else:
                query["property_record_id"] = {"$in": search_property_ids}
        else:
            # No matching properties found
            return {
                "submissions": [],
                "total": 0,
                "page": page,
                "pages": 0
            }
    
    # If colony filter is provided, first get property IDs in that colony
    property_ids_in_colony = None
    if colony and colony.strip():
        properties_in_colony = await db.properties.find(
            {"ward": colony}, 
            {"id": 1, "_id": 0}
        ).to_list(None)
        property_ids_in_colony = [p["id"] for p in properties_in_colony]
        if property_ids_in_colony:
            if "property_record_id" in query:
                # Intersect with existing filter
                existing_ids = set(query["property_record_id"].get("$in", []))
                query["property_record_id"]["$in"] = list(existing_ids & set(property_ids_in_colony)) if existing_ids else property_ids_in_colony
            else:
                query["property_record_id"] = {"$in": property_ids_in_colony}
        else:
            # No properties in this colony, return empty
            return {
                "submissions": [],
                "total": 0,
                "page": page,
                "pages": 0
            }
    
    skip = (page - 1) * limit
    total = await db.submissions.count_documents(query)
    submissions = await db.submissions.find(query, {"_id": 0}).sort("submitted_at", -1).skip(skip).limit(limit).to_list(limit)
    
    # Enrich with property details
    for sub in submissions:
        if sub.get("property_record_id"):
            prop = await db.properties.find_one({"id": sub["property_record_id"]}, {"_id": 0})
            if prop:
                sub["property_owner_name"] = prop.get("owner_name", "")
                sub["property_mobile"] = prop.get("mobile", "")
                sub["property_address"] = prop.get("address", "")
                sub["property_amount"] = prop.get("amount", "")
                sub["property_ward"] = prop.get("ward", "")
                sub["colony"] = prop.get("colony") or prop.get("ward", "")
                sub["total_area"] = prop.get("total_area", "")
                sub["category"] = prop.get("category", "")
                # Add serial number info
                sub["serial_number"] = prop.get("serial_number", 0)
                sub["bill_sr_no"] = prop.get("bill_sr_no", "")
                sub["property_serial_number"] = prop.get("serial_number", 0)
                sub["property_serial_na"] = prop.get("serial_na", False)
                sub["property_bill_sr_no"] = prop.get("bill_sr_no", "N/A")
    
    return {
        "submissions": submissions,
        "total": total,
        "page": page,
        "pages": (total + limit - 1) // limit
    }

@api_router.post("/admin/submissions/approve")
async def approve_reject_submission(data: SubmissionApproval, current_user: dict = Depends(get_current_user)):
    if current_user["role"] != "ADMIN":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    submission = await db.submissions.find_one({"id": data.submission_id}, {"_id": 0})
    if not submission:
        raise HTTPException(status_code=404, detail="Submission not found")
    
    if data.action == "REJECT" and not data.remarks:
        raise HTTPException(status_code=400, detail="Remarks are required for rejection")
    
    new_status = "Approved" if data.action == "APPROVE" else "Rejected"
    
    update_data = {
        "status": new_status,
        "reviewed_by": current_user["id"],
        "reviewed_at": datetime.now(timezone.utc).isoformat()
    }
    
    if data.remarks:
        update_data["review_remarks"] = data.remarks
    
    await db.submissions.update_one(
        {"id": data.submission_id},
        {"$set": update_data}
    )
    
    # Update property status and include rejection remarks
    prop_status = "Approved" if data.action == "APPROVE" else "Rejected"
    prop_update = {"status": prop_status}
    
    # Lock the property if approved (prevents re-submission)
    if data.action == "APPROVE":
        prop_update["locked"] = True
        prop_update["locked_at"] = datetime.now(timezone.utc).isoformat()
        prop_update["locked_by"] = current_user["id"]
    
    if data.action == "REJECT" and data.remarks:
        prop_update["reject_remarks"] = data.remarks
        # Unlock property on rejection so surveyor can re-submit
        prop_update["locked"] = False
    
    await db.properties.update_one(
        {"id": submission["property_record_id"]},
        {"$set": prop_update}
    )
    
    return {"message": f"Submission {new_status.lower()}"}

@api_router.put("/admin/submissions/{submission_id}")
async def edit_submission(
    submission_id: str,
    request: Request,
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] not in SUBMISSION_EDIT_ROLES:
        raise HTTPException(status_code=403, detail="Admin access required for editing submissions")
    
    data = await request.json()
    
    update_data = {}
    allowed_fields = [
        "receiver_name", "receiver_mobile", "relation",
        "correct_colony_name", "self_satisfied",
        "remarks", "latitude", "longitude",
        "new_owner_name", "new_mobile",
        "special_condition"  # Added for admin edit
    ]
    
    for field in allowed_fields:
        if field in data:
            value = data[field]
            # Convert latitude/longitude to float if provided
            if field in ["latitude", "longitude"] and value:
                try:
                    update_data[field] = float(value)
                except ValueError:
                    pass
            else:
                update_data[field] = value
    
    # Handle photos update
    if "photos" in data:
        update_data["photos"] = data["photos"]
    
    update_data["edited_by"] = current_user["id"]
    update_data["edited_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.submissions.update_one(
        {"id": submission_id},
        {"$set": update_data}
    )
    
    return {"message": "Submission updated"}

@api_router.post("/admin/submissions/upload-photo")
async def upload_submission_photo(
    file: UploadFile = File(...),
    submission_id: str = Form(...),
    photo_type: str = Form("HOUSE"),
    current_user: dict = Depends(get_current_user)
):
    """Upload a new photo to a submission (admin only) - saves to GridFS"""
    if current_user["role"] not in SUBMISSION_EDIT_ROLES:
        raise HTTPException(status_code=403, detail="Admin access required for editing submissions")
    
    # Read file content
    content = await file.read()
    
    # Save to GridFS
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{submission_id}_{photo_type.lower()}_{timestamp}.jpg"
    file_id = await save_file_to_gridfs(content, filename, file.content_type or "image/jpeg")
    
    file_url = f"/api/file/{file_id}"
    
    # Add photo to submission
    photo_data = {
        "file_url": file_url,
        "file_id": file_id,
        "photo_type": photo_type
    }
    
    await db.submissions.update_one(
        {"id": submission_id},
        {"$push": {"photos": photo_data}}
    )
    
    return {"message": "Photo uploaded", "file_url": file_url, "file_id": file_id}

@api_router.put("/admin/properties/{property_id}")
async def edit_property(
    property_id: str,
    request: Request,
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] != "ADMIN":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    data = await request.json()
    
    update_data = {}
    allowed_fields = [
        "property_id", "owner_name", "mobile", "address", "amount", "ward"
    ]
    
    for field in allowed_fields:
        if field in data and data[field]:
            update_data[field] = data[field]
    
    update_data["edited_by"] = current_user["id"]
    update_data["edited_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.properties.update_one(
        {"id": property_id},
        {"$set": update_data}
    )
    
    return {"message": "Property updated"}

# ============== EXPORT ROUTES ==============

@api_router.get("/admin/export")
async def export_data(
    batch_id: Optional[str] = None,
    employee_id: Optional[str] = None,
    status: Optional[str] = "Approved",  # Default to Approved
    colony: Optional[str] = None,  # Colony filter
    date_from: Optional[str] = None,  # Date filter
    date_to: Optional[str] = None,  # Date filter
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] not in EXPORT_ROLES:
        raise HTTPException(status_code=403, detail="Export access required (Admin or MC Officer)")
    
    # Build query for properties
    prop_query = {}
    if batch_id and batch_id.strip():
        prop_query["batch_id"] = batch_id
    if employee_id and employee_id.strip():
        prop_query["assigned_employee_id"] = employee_id
    if colony and colony.strip():
        prop_query["ward"] = colony
    
    # Build submission query for status and date filters
    submission_query = {}
    if status and status.strip():
        submission_query["status"] = status
    if date_from:
        submission_query["submitted_at"] = {"$gte": date_from}
    if date_to:
        if "submitted_at" in submission_query:
            submission_query["submitted_at"]["$lte"] = date_to
        else:
            submission_query["submitted_at"] = {"$lte": date_to}
    
    # If we have submission filters, get property IDs from matching submissions
    if submission_query:
        submissions = await db.submissions.find(submission_query, {"property_record_id": 1, "_id": 0}).to_list(100000)
        property_ids = [s["property_record_id"] for s in submissions]
        
        if not property_ids:
            # No submissions match, return empty Excel
            wb = Workbook()
            ws = wb.active
            ws.title = "Approved Survey Data"
            ws.cell(row=1, column=1, value="No submissions found matching the filters")
            export_path = UPLOAD_DIR / f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            wb.save(export_path)
            return FileResponse(
                path=str(export_path),
                filename=f"approved_survey_export_{datetime.now().strftime('%Y%m%d')}.xlsx",
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        
        if "id" in prop_query:
            # Intersect with existing property_ids filter
            existing_ids = set(prop_query["id"]["$in"]) if isinstance(prop_query["id"], dict) else {prop_query["id"]}
            prop_query["id"] = {"$in": list(existing_ids.intersection(set(property_ids)))}
        else:
            prop_query["id"] = {"$in": property_ids}
    
    properties = await db.properties.find(prop_query, {"_id": 0}).to_list(100000)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Property Survey Data"
    
    header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    headers = [
        "Property ID", "Owner Name", "Mobile", "Address", "Total Area", "Amount", "Ward",
        "Assigned Employee", "Status", "New Owner Name", "New Mobile", "Receiver Name",
        "Relation", "Old Property ID", "Family ID", "Aadhar Number", "Ward Number",
        "GPS Latitude", "GPS Longitude", "Submission Date", "Signature URL", "Photo URLs",
        "Approval Status", "Review Remarks"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    for row_idx, prop in enumerate(properties, 2):
        submission = await db.submissions.find_one(
            {"property_record_id": prop["id"]}, 
            {"_id": 0}
        )
        
        ws.cell(row=row_idx, column=1, value=prop.get("property_id", ""))
        ws.cell(row=row_idx, column=2, value=prop.get("owner_name", ""))
        ws.cell(row=row_idx, column=3, value=prop.get("mobile", ""))
        ws.cell(row=row_idx, column=4, value=prop.get("address", ""))
        ws.cell(row=row_idx, column=5, value=prop.get("total_area", ""))
        ws.cell(row=row_idx, column=6, value=prop.get("amount", ""))
        ws.cell(row=row_idx, column=7, value=prop.get("ward", ""))
        ws.cell(row=row_idx, column=8, value=prop.get("assigned_employee_name", ""))
        ws.cell(row=row_idx, column=9, value=prop.get("status", ""))
        
        if submission:
            ws.cell(row=row_idx, column=10, value=submission.get("new_owner_name", ""))
            ws.cell(row=row_idx, column=11, value=submission.get("new_mobile", ""))
            ws.cell(row=row_idx, column=12, value=submission.get("receiver_name", ""))
            ws.cell(row=row_idx, column=13, value=submission.get("relation", ""))
            ws.cell(row=row_idx, column=14, value=submission.get("old_property_id", ""))
            ws.cell(row=row_idx, column=15, value=submission.get("family_id", ""))
            ws.cell(row=row_idx, column=16, value=submission.get("aadhar_number", ""))
            ws.cell(row=row_idx, column=17, value=submission.get("ward_number", ""))
            ws.cell(row=row_idx, column=18, value=submission.get("latitude", ""))
            ws.cell(row=row_idx, column=19, value=submission.get("longitude", ""))
            ws.cell(row=row_idx, column=20, value=submission.get("submitted_at", ""))
            ws.cell(row=row_idx, column=21, value=submission.get("signature_url", ""))
            photos = submission.get("photos", [])
            photo_urls = ", ".join([p.get("file_url", "") for p in photos])
            ws.cell(row=row_idx, column=22, value=photo_urls)
            ws.cell(row=row_idx, column=23, value=submission.get("status", "Pending"))
            ws.cell(row=row_idx, column=24, value=submission.get("review_remarks", ""))
    
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    export_path = UPLOAD_DIR / f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(export_path)
    
    return FileResponse(
        path=str(export_path),
        filename=f"property_survey_export_{datetime.now().strftime('%Y%m%d')}.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# Helper function to add watermark to photo
def add_watermark_to_photo(photo_path, latitude, longitude, submitted_at):
    try:
        img = PILImage.open(photo_path)
        draw = ImageDraw.Draw(img)
        
        if isinstance(submitted_at, str):
            try:
                dt = datetime.fromisoformat(submitted_at.replace('Z', '+00:00'))
            except:
                dt = datetime.now()
        else:
            dt = submitted_at or datetime.now()
        
        date_str = dt.strftime("%d/%m/%Y")
        time_str = dt.strftime("%I:%M:%S %p")
        
        watermark_lines = [
            f"Date: {date_str}",
            f"Time: {time_str}",
            f"Lat: {latitude:.6f}" if latitude else "Lat: N/A",
            f"Long: {longitude:.6f}" if longitude else "Long: N/A"
        ]
        
        font_size = max(16, min(img.width, img.height) // 25)
        try:
            font = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", font_size)
        except:
            font = ImageFont.load_default()
        
        padding = font_size // 2
        line_height = font_size + 5
        
        max_text_width = max([draw.textlength(line, font=font) for line in watermark_lines])
        box_width = int(max_text_width + padding * 2)
        box_height = line_height * len(watermark_lines) + padding * 2
        
        box_x = padding
        box_y = img.height - box_height - padding
        
        overlay = PILImage.new('RGBA', img.size, (0, 0, 0, 0))
        overlay_draw = ImageDraw.Draw(overlay)
        overlay_draw.rectangle(
            [box_x, box_y, box_x + box_width, box_y + box_height],
            fill=(0, 0, 0, 180)
        )
        
        if img.mode != 'RGBA':
            img = img.convert('RGBA')
        
        img = PILImage.alpha_composite(img, overlay)
        draw = ImageDraw.Draw(img)
        
        for i, line in enumerate(watermark_lines):
            draw.text(
                (box_x + padding, box_y + padding + i * line_height),
                line,
                font=font,
                fill=(255, 255, 255, 255)
            )
        
        img = img.convert('RGB')
        
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.jpg')
        img.save(temp_file.name, 'JPEG', quality=90)
        return temp_file.name
    except Exception as e:
        logger.error(f"Error adding watermark: {e}")
        return photo_path

@api_router.get("/admin/export-pdf")
async def export_pdf(
    batch_id: Optional[str] = None,
    employee_id: Optional[str] = None,
    status: Optional[str] = "Approved",  # Default to Approved
    colony: Optional[str] = None,  # Colony filter
    date_from: Optional[str] = None,  # Date filter
    date_to: Optional[str] = None,  # Date filter
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] not in EXPORT_ROLES:
        raise HTTPException(status_code=403, detail="Export access required (Admin or MC Officer)")
    
    # Build submission query with status and date filters
    submission_query = {"status": status if status and status.strip() else "Approved"}
    if date_from:
        submission_query["submitted_at"] = {"$gte": date_from}
    if date_to:
        if "submitted_at" in submission_query:
            submission_query["submitted_at"]["$lte"] = date_to
        else:
            submission_query["submitted_at"] = {"$lte": date_to}
    
    submissions = await db.submissions.find(submission_query, {"_id": 0}).to_list(10000)
    
    if not submissions:
        # Return empty PDF with message
        pdf_path = UPLOAD_DIR / f"survey_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        doc = SimpleDocTemplate(str(pdf_path), pagesize=A4)
        styles = getSampleStyleSheet()
        story = [Paragraph("No submissions found matching the filters", styles['Heading1'])]
        doc.build(story)
        return FileResponse(
            path=str(pdf_path),
            filename=f"approved_survey_report_{datetime.now().strftime('%Y%m%d')}.pdf",
            media_type="application/pdf"
        )
    
    # Get property IDs from submissions
    property_ids = [s["property_record_id"] for s in submissions]
    
    # Build property query
    prop_query = {"id": {"$in": property_ids}}
    if batch_id and batch_id.strip():
        prop_query["batch_id"] = batch_id
    if employee_id and employee_id.strip():
        prop_query["assigned_employee_id"] = employee_id
    if colony and colony.strip():
        prop_query["ward"] = colony
    
    properties = await db.properties.find(prop_query, {"_id": 0}).to_list(10000)
    
    pdf_path = UPLOAD_DIR / f"survey_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    doc = SimpleDocTemplate(
        str(pdf_path),
        pagesize=A4,
        rightMargin=20*mm,
        leftMargin=20*mm,
        topMargin=20*mm,
        bottomMargin=20*mm
    )
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=18, spaceAfter=20, alignment=TA_CENTER, textColor=colors.HexColor('#0f172a'))
    heading_style = ParagraphStyle('CustomHeading', parent=styles['Heading2'], fontSize=14, spaceAfter=10, textColor=colors.HexColor('#1e40af'))
    normal_style = ParagraphStyle('CustomNormal', parent=styles['Normal'], fontSize=10, spaceAfter=5)
    
    story = []
    
    story.append(Paragraph("NSTU Property Tax Survey Report", title_style))
    story.append(Paragraph(f"Generated on: {datetime.now().strftime('%d/%m/%Y %I:%M %p')}", normal_style))
    story.append(Spacer(1, 20))
    
    for prop in properties:
        submission = await db.submissions.find_one({"property_record_id": prop["id"]}, {"_id": 0})
        
        if not submission:
            continue
        
        story.append(Paragraph(f"Property ID: {prop.get('property_id', 'N/A')}", heading_style))
        
        prop_data = [
            ["Owner Name", prop.get("owner_name", "N/A")],
            ["Mobile", prop.get("mobile", "N/A")],
            ["Address", prop.get("address", "N/A")],
            ["Ward", prop.get("ward", "N/A")],
            ["Amount", prop.get("amount", "N/A")],
        ]
        
        prop_table = Table(prop_data, colWidths=[80*mm, 90*mm])
        prop_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f1f5f9')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#0f172a')),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('PADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        story.append(prop_table)
        story.append(Spacer(1, 10))
        
        story.append(Paragraph("Survey Information", heading_style))
        survey_data = [
            ["New Owner Name", submission.get("new_owner_name", "N/A")],
            ["New Mobile", submission.get("new_mobile", "N/A")],
            ["Receiver Name", submission.get("receiver_name", "N/A")],
            ["Relation", submission.get("relation", "N/A")],
            ["Old Property ID", submission.get("old_property_id", "N/A")],
            ["Family ID", submission.get("family_id", "N/A")],
            ["Aadhar Number", submission.get("aadhar_number", "N/A")],
            ["Ward Number", submission.get("ward_number", "N/A")],
            ["Submitted By", submission.get("employee_name", "N/A")],
            ["Submitted At", submission.get("submitted_at", "N/A")],
            ["GPS Latitude", str(submission.get("latitude", "N/A"))],
            ["GPS Longitude", str(submission.get("longitude", "N/A"))],
            ["Status", submission.get("status", "Pending")],
        ]
        
        if submission.get("remarks"):
            survey_data.append(["Remarks", submission.get("remarks")])
        
        survey_table = Table(survey_data, colWidths=[80*mm, 90*mm])
        survey_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f1f5f9')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#0f172a')),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('PADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        story.append(survey_table)
        story.append(Spacer(1, 15))
        
        photos = submission.get("photos", [])
        if photos:
            story.append(Paragraph("Photo Evidence (with GPS & Timestamp)", heading_style))
            
            for photo in photos:
                photo_url = photo.get("file_url", "")
                photo_type = photo.get("photo_type", "PHOTO")
                file_id = photo.get("file_id")
                
                temp_photo_path = None
                
                # Handle GridFS files (new format)
                if file_id or photo_url.startswith("/api/file/"):
                    try:
                        grid_file_id = file_id or photo_url.replace("/api/file/", "")
                        content, filename, _ = await get_file_from_gridfs(grid_file_id)
                        if content:
                            temp_photo_path = UPLOAD_DIR / f"temp_pdf_{uuid.uuid4()}{Path(filename).suffix}"
                            async with aiofiles.open(temp_photo_path, 'wb') as f:
                                await f.write(content)
                    except Exception as e:
                        logger.error(f"Error fetching GridFS photo: {e}")
                
                # Handle legacy local files
                elif photo_url.startswith("/api/uploads/"):
                    filename = photo_url.replace("/api/uploads/", "")
                    photo_path = UPLOAD_DIR / filename
                    if photo_path.exists():
                        temp_photo_path = photo_path
                
                if temp_photo_path and temp_photo_path.exists():
                    watermarked_path = add_watermark_to_photo(
                        str(temp_photo_path),
                        submission.get("latitude"),
                        submission.get("longitude"),
                        submission.get("submitted_at")
                    )
                    
                    try:
                        img = RLImage(watermarked_path, width=80*mm, height=60*mm)
                        story.append(Paragraph(f"<b>{photo_type}</b>", normal_style))
                        story.append(img)
                        story.append(Spacer(1, 10))
                    except Exception as e:
                        logger.error(f"Error adding photo to PDF: {e}")
                    finally:
                        # Cleanup temp files from GridFS
                        if file_id or photo_url.startswith("/api/file/"):
                            try:
                                if temp_photo_path.exists():
                                    temp_photo_path.unlink()
                            except:
                                pass
        
        signature_url = submission.get("signature_url")
        if signature_url:
            story.append(Paragraph("Property Holder Signature", heading_style))
            
            temp_sig_path = None
            
            # Handle GridFS signature (new format)
            if signature_url.startswith("/api/file/"):
                try:
                    sig_file_id = signature_url.replace("/api/file/", "")
                    content, filename, _ = await get_file_from_gridfs(sig_file_id)
                    if content:
                        temp_sig_path = UPLOAD_DIR / f"temp_sig_{uuid.uuid4()}.png"
                        async with aiofiles.open(temp_sig_path, 'wb') as f:
                            await f.write(content)
                except Exception as e:
                    logger.error(f"Error fetching GridFS signature: {e}")
            
            # Handle legacy local files
            elif signature_url.startswith("/api/uploads/"):
                sig_filename = signature_url.replace("/api/uploads/", "")
                sig_path = UPLOAD_DIR / sig_filename
                if sig_path.exists():
                    temp_sig_path = sig_path
            
            if temp_sig_path and temp_sig_path.exists():
                try:
                    sig_img = RLImage(str(temp_sig_path), width=60*mm, height=30*mm)
                    sig_table = Table([[sig_img]], colWidths=[170*mm])
                    sig_table.setStyle(TableStyle([
                        ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#e2e8f0')),
                        ('BACKGROUND', (0, 0), (-1, -1), colors.white),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('PADDING', (0, 0), (-1, -1), 10),
                    ]))
                    story.append(sig_table)
                except Exception as e:
                    logger.error(f"Error adding signature to PDF: {e}")
                finally:
                    # Cleanup temp files from GridFS
                    if signature_url.startswith("/api/file/"):
                        try:
                            if temp_sig_path.exists():
                                temp_sig_path.unlink()
                        except:
                            pass
        
        story.append(PageBreak())
    
    doc.build(story)
    
    return FileResponse(
        path=str(pdf_path),
        filename=f"property_survey_report_{datetime.now().strftime('%Y%m%d')}.pdf",
        media_type="application/pdf"
    )

# ============== EMPLOYEE ROUTES ==============

@api_router.get("/employee/properties")
async def get_employee_properties(
    search: Optional[str] = None,
    status: Optional[str] = None,
    page: int = 1,
    limit: int = 50,  # Increased default for map view
    current_user: dict = Depends(get_current_user)
):
    # Check both single assigned_employee_id and array assigned_employee_ids
    query = {
        "$or": [
            {"assigned_employee_id": current_user["id"]},
            {"assigned_employee_ids": current_user["id"]}
        ]
    }
    if status and status.strip():
        query["status"] = status
    if search:
        query["$and"] = query.get("$and", []) + [{
            "$or": [
                {"property_id": {"$regex": search, "$options": "i"}},
                {"owner_name": {"$regex": search, "$options": "i"}},
                {"mobile": {"$regex": search, "$options": "i"}}
            ]
        }]
    
    # Optimized projection - only return fields needed for map and list view
    projection = {
        "_id": 0,
        "id": 1,
        "property_id": 1,
        "owner_name": 1,
        "mobile": 1,
        "address": 1,
        "colony": 1,
        "ward": 1,
        "latitude": 1,
        "longitude": 1,
        "status": 1,
        "serial_number": 1,
        "bill_sr_no": 1,
        "amount": 1,
        "category": 1,
        "total_area": 1
    }
    
    skip = (page - 1) * limit
    total = await db.properties.count_documents(query)
    
    # Use sort for consistent ordering - pending first, then by serial number
    properties = await db.properties.find(query, projection).sort([
        ("status", 1),  # Pending first
        ("serial_number", 1)
    ]).skip(skip).limit(limit).to_list(limit)
    
    return {
        "properties": properties,
        "total": total,
        "page": page,
        "pages": (total + limit - 1) // limit
    }

@api_router.get("/employee/property/{property_id}")
async def get_property_detail(property_id: str, current_user: dict = Depends(get_current_user)):
    prop = await db.properties.find_one({"id": property_id}, {"_id": 0})
    if not prop:
        raise HTTPException(status_code=404, detail="Property not found")
    
    # Check if employee is assigned (either single or in array)
    is_assigned = (
        prop.get("assigned_employee_id") == current_user["id"] or
        current_user["id"] in (prop.get("assigned_employee_ids") or [])
    )
    if current_user["role"] != "ADMIN" and not is_assigned:
        raise HTTPException(status_code=403, detail="Access denied")
    
    submission = await db.submissions.find_one({"property_record_id": property_id}, {"_id": 0})
    
    return {
        "property": prop,
        "submission": submission
    }

@api_router.post("/employee/submit/{property_id}")
async def submit_survey(
    property_id: str,
    # Survey fields - Simplified as per requirements
    receiver_name: str = Form(""),
    receiver_mobile: str = Form(""),
    relation: str = Form(""),
    correct_colony_name: str = Form(None),
    remarks: str = Form(None),
    self_satisfied: str = Form(""),
    special_condition: str = Form(None),  # NEW: 'house_locked' or 'owner_denied'
    # Self Certification fields
    self_cert_status: str = Form(None),  # 'done', 'later', 'deny', 'already_certified'
    self_cert_mobile: str = Form(None),
    self_cert_otp: str = Form(None),
    latitude: float = Form(...),
    longitude: float = Form(...),
    house_photo: UploadFile = File(None),  # Now optional
    gate_photo: UploadFile = File(None),   # Now optional
    signature: UploadFile = File(None),    # Now optional
    extra_photos: List[UploadFile] = File(default=[]),
    authorization: str = Form(...),
    # Legacy fields - keep for backward compatibility
    new_owner_name: str = Form(None),
    new_mobile: str = Form(None),
    old_property_id: str = Form(None),
    family_id: str = Form(None),
    aadhar_number: str = Form(None),
    ward_number: str = Form(None)
):
    current_user = await get_current_user(authorization)
    
    prop = await db.properties.find_one({"id": property_id}, {"_id": 0})
    if not prop:
        raise HTTPException(status_code=404, detail="Property not found")
    
    # Check if property is locked (survey already completed and approved)
    if prop.get("locked") == True:
        raise HTTPException(status_code=403, detail="This property is locked. Survey already completed.")
    
    # Check if property status is Completed or Approved - prevent re-submission
    if prop.get("status") in ["Completed", "Approved"]:
        raise HTTPException(status_code=403, detail="This property survey is already completed. Cannot re-submit.")
    
    # Check if employee is assigned (either single or in array)
    is_assigned = (
        prop.get("assigned_employee_id") == current_user["id"] or
        current_user["id"] in (prop.get("assigned_employee_ids") or [])
    )
    if current_user["role"] != "ADMIN" and not is_assigned:
        raise HTTPException(status_code=403, detail="Access denied")
    
    # Check if special condition allows skipping required fields
    is_special_condition = special_condition in ['house_locked', 'owner_denied']
    
    # Validate required fields only if not special condition
    if not is_special_condition:
        if not receiver_name or not relation or not receiver_mobile or not self_satisfied:
            raise HTTPException(status_code=400, detail="Receiver name, mobile, relation and satisfaction status are required")
        if not house_photo:
            raise HTTPException(status_code=400, detail="Property photo is required")
    
    photos = []
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    signature_url = None
    
    # House photo - SAVE TO GRIDFS
    if house_photo and house_photo.filename:
        content = await house_photo.read()
        house_filename = f"{property_id}_house_{timestamp}{Path(house_photo.filename).suffix}"
        file_id = await save_file_to_gridfs(content, house_filename, house_photo.content_type or "image/jpeg")
        photos.append({"photo_type": "HOUSE", "file_url": f"/api/file/{file_id}", "file_id": file_id})
    
    # Gate photo - SAVE TO GRIDFS
    if gate_photo and gate_photo.filename:
        content = await gate_photo.read()
        gate_filename = f"{property_id}_gate_{timestamp}{Path(gate_photo.filename).suffix}"
        file_id = await save_file_to_gridfs(content, gate_filename, gate_photo.content_type or "image/jpeg")
        photos.append({"photo_type": "GATE", "file_url": f"/api/file/{file_id}", "file_id": file_id})
    
    # Signature - SAVE TO GRIDFS
    if signature and signature.filename:
        content = await signature.read()
        signature_filename = f"{property_id}_signature_{timestamp}.png"
        file_id = await save_file_to_gridfs(content, signature_filename, "image/png")
        signature_url = f"/api/file/{file_id}"
    
    # Extra photos - SAVE TO GRIDFS
    for idx, photo in enumerate(extra_photos):
        if photo.filename:
            content = await photo.read()
            extra_filename = f"{property_id}_extra{idx}_{timestamp}{Path(photo.filename).suffix}"
            file_id = await save_file_to_gridfs(content, extra_filename, photo.content_type or "image/jpeg")
            photos.append({"photo_type": "EXTRA", "file_url": f"/api/file/{file_id}", "file_id": file_id})
    
    # Set receiver name based on special condition if empty
    final_receiver_name = receiver_name
    if is_special_condition and not receiver_name:
        final_receiver_name = "House Locked" if special_condition == 'house_locked' else "Owner Denied"
    
    # Create submission with new fields
    submission_doc = {
        "id": str(uuid.uuid4()),
        "property_record_id": property_id,
        "property_id": prop["property_id"],
        "batch_id": prop["batch_id"],
        "employee_id": current_user["id"],
        "employee_name": current_user["name"],
        # Survey fields - NEW simplified structure
        "receiver_name": final_receiver_name,
        "receiver_mobile": receiver_mobile,
        "relation": relation or ("N/A" if is_special_condition else ""),
        "correct_colony_name": correct_colony_name,
        "remarks": remarks,
        "self_satisfied": self_satisfied or ("N/A" if is_special_condition else "yes"),
        "special_condition": special_condition,  # NEW field
        # Self Certification data
        "self_cert_status": self_cert_status,  # done, later, deny, already_certified
        "self_cert_mobile": self_cert_mobile if self_cert_status == 'done' else None,
        "self_cert_otp": self_cert_otp if self_cert_status == 'done' else None,
        "self_cert_verified": True if self_cert_status == 'done' and self_cert_otp else False,
        "latitude": latitude,
        "longitude": longitude,
        "submitted_at": datetime.now(timezone.utc).isoformat(),
        "photos": photos,
        "signature_url": signature_url,
        "status": "Pending",
        # Legacy fields for backward compat
        "new_owner_name": new_owner_name or prop.get("owner_name"),
        "new_mobile": new_mobile or prop.get("mobile"),
        "old_property_id": old_property_id,
        "family_id": family_id,
        "aadhar_number": aadhar_number,
        "ward_number": ward_number
    }
    
    # Check if submission already exists
    existing = await db.submissions.find_one({"property_record_id": property_id})
    if existing:
        await db.submissions.update_one(
            {"property_record_id": property_id},
            {"$set": submission_doc}
        )
    else:
        await db.submissions.insert_one(submission_doc)
    
    # Update property status to In Progress (until approved)
    await db.properties.update_one(
        {"id": property_id},
        {"$set": {"status": "In Progress"}}
    )
    
    # Clear map cache so other users see updated status
    await clear_map_cache()
    
    return {"message": "Survey submitted successfully", "submission_id": submission_doc["id"]}

@api_router.post("/employee/reject/{property_id}")
async def reject_property(property_id: str, remarks: str = Form(...), authorization: str = Form(...)):
    current_user = await get_current_user(authorization)
    
    prop = await db.properties.find_one({"id": property_id}, {"_id": 0})
    if not prop:
        raise HTTPException(status_code=404, detail="Property not found")
    
    # Check if employee is assigned (either single or in array)
    is_assigned = (
        prop.get("assigned_employee_id") == current_user["id"] or
        current_user["id"] in (prop.get("assigned_employee_ids") or [])
    )
    if current_user["role"] != "ADMIN" and not is_assigned:
        raise HTTPException(status_code=403, detail="Access denied")
    
    await db.properties.update_one(
        {"id": property_id},
        {"$set": {"status": "Rejected", "reject_remarks": remarks}}
    )
    
    return {"message": "Property rejected"}

@api_router.get("/employee/progress")
async def get_employee_own_progress(current_user: dict = Depends(get_current_user)):
    today_start = get_today_start().isoformat()
    
    # Query for properties assigned to this employee (single or in array)
    assign_query = {
        "$or": [
            {"assigned_employee_id": current_user["id"]},
            {"assigned_employee_ids": current_user["id"]}
        ]
    }
    
    total = await db.properties.count_documents(assign_query)
    completed = await db.properties.count_documents({
        **assign_query,
        "status": "Completed"
    })
    pending = await db.properties.count_documents({
        **assign_query,
        "status": "Pending"
    })
    rejected = await db.properties.count_documents({
        **assign_query,
        "status": "Rejected"
    })
    in_progress = await db.properties.count_documents({
        **assign_query,
        "status": "In Progress"
    })
    
    # Today's completed
    today_completed = await db.submissions.count_documents({
        "employee_id": current_user["id"],
        "submitted_at": {"$gte": today_start}
    })
    
    # Total completed (all time)
    total_completed = await db.submissions.count_documents({
        "employee_id": current_user["id"]
    })
    
    return {
        "total_assigned": total,
        "completed": completed,
        "pending": pending,
        "rejected": rejected,
        "in_progress": in_progress,
        "today_completed": today_completed,
        "total_completed": total_completed
    }

# ============== ATTENDANCE ROUTES ==============

@api_router.get("/employee/attendance/today")
async def check_today_attendance(current_user: dict = Depends(get_current_user)):
    """Check if employee has marked attendance today"""
    today_date = get_today_start().strftime("%Y-%m-%d")
    
    attendance = await db.attendance.find_one({
        "employee_id": current_user["id"],
        "date": today_date
    }, {"_id": 0})
    
    return {
        "marked": attendance is not None,
        "has_attendance": attendance is not None,
        "attendance": attendance
    }

@api_router.post("/employee/attendance")
async def mark_attendance(
    selfie: UploadFile = File(...),
    latitude: float = Form(...),
    longitude: float = Form(...),
    authorization: str = Form(...)
):
    """Mark one-time daily attendance with selfie"""
    current_user = await get_current_user(authorization)
    today_date = get_today_start().strftime("%Y-%m-%d")
    
    # Check if already marked
    existing = await db.attendance.find_one({
        "employee_id": current_user["id"],
        "date": today_date
    })
    
    if existing:
        raise HTTPException(status_code=400, detail="Attendance already marked for today")
    
    # Save selfie TO GRIDFS
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    selfie_filename = f"attendance_{current_user['id']}_{timestamp}{Path(selfie.filename).suffix}"
    content = await selfie.read()
    file_id = await save_file_to_gridfs(content, selfie_filename, selfie.content_type or "image/jpeg")
    selfie_url = f"/api/file/{file_id}"
    
    # Create attendance record
    attendance_doc = {
        "id": str(uuid.uuid4()),
        "employee_id": current_user["id"],
        "employee_name": current_user["name"],
        "date": today_date,
        "marked_at": datetime.now(timezone.utc).isoformat(),
        "selfie_url": selfie_url,
        "selfie_file_id": file_id,
        "latitude": latitude,
        "longitude": longitude
    }
    
    await db.attendance.insert_one(attendance_doc)
    
    return {
        "message": "Attendance marked successfully",
        "attendance_id": attendance_doc["id"],
        "marked_at": attendance_doc["marked_at"]
    }

@api_router.get("/admin/attendance")
async def get_attendance_records(
    date: Optional[str] = None,
    employee_id: Optional[str] = None,
    page: int = 1,
    limit: int = 50,
    current_user: dict = Depends(get_current_user)
):
    """Get attendance records (admin/supervisor only)"""
    if current_user["role"] not in ADMIN_VIEW_ROLES:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    query = {}
    if date:
        query["date"] = date
    if employee_id and employee_id.strip():
        query["employee_id"] = employee_id
    
    skip = (page - 1) * limit
    total = await db.attendance.count_documents(query)
    records = await db.attendance.find(query, {"_id": 0}).sort("marked_at", -1).skip(skip).limit(limit).to_list(limit)
    
    return {
        "attendance": records,
        "total": total,
        "page": page,
        "pages": (total + limit - 1) // limit
    }

# ============== PDF BILL PROCESSING ==============

# Helper function to extract BillSrNo from page using block-based extraction
def extract_bill_sr_no_from_page(page) -> str:
    """Extract BillSrNo from PDF page using position-aware extraction"""
    try:
        blocks = page.get_text("dict")["blocks"]
        billsr_y = None
        billsr_x = None
        
        # First find the BillSrNo label position
        for block in blocks:
            if "lines" in block:
                for line in block["lines"]:
                    for span in line["spans"]:
                        text = span["text"].strip().lower()
                        if 'billsrno' in text or 'bill sr no' in text:
                            bbox = span["bbox"]
                            billsr_y = bbox[1]  # y position
                            billsr_x = bbox[2]  # x position (right edge of label)
                            # Check if number is on same line after ":"
                            full_text = span["text"]
                            match = re.search(r'[:\s]+(\d+)\s*$', full_text)
                            if match:
                                return match.group(1)
        
        # If we found BillSrNo label, look for a number nearby
        if billsr_y is not None:
            candidates = []
            for block in blocks:
                if "lines" in block:
                    for line in block["lines"]:
                        for span in line["spans"]:
                            text = span["text"].strip()
                            bbox = span["bbox"]
                            # Look for standalone numbers in similar y position (within 20 pixels)
                            if text.isdigit() and abs(bbox[1] - billsr_y) < 20:
                                # Prefer numbers to the right of the label
                                if billsr_x is None or bbox[0] >= billsr_x - 50:
                                    candidates.append((abs(bbox[1] - billsr_y), text))
            
            if candidates:
                # Return the closest number
                candidates.sort(key=lambda x: x[0])
                return candidates[0][1]
        
        return ""
    except Exception:
        return ""

# Helper function to extract bill data from PDF text
def extract_bill_data(text: str, page_num: int, page=None) -> dict:
    """Extract structured bill data from PDF page text"""
    
    def find_value(patterns, text, default=""):
        for pattern in patterns:
            match = re.search(pattern, text, re.IGNORECASE | re.MULTILINE)
            if match:
                return match.group(1).strip()
        return default
    
    # Extract coordinates (latitude : longitude format)
    coords_match = re.search(r'(\d+\.\d+)\s*:\s*(\d+\.\d+)', text)
    latitude = float(coords_match.group(1)) if coords_match else None
    longitude = float(coords_match.group(2)) if coords_match else None
    
    # Try to extract BillSrNo using block-based extraction first
    bill_sr_no = ""
    if page is not None:
        bill_sr_no = extract_bill_sr_no_from_page(page)
    
    # Fallback to regex if block extraction failed
    if not bill_sr_no:
        bill_sr_no = find_value([
            r'BillSrNo\.?\s*[:\s]*\s*(\d+)',      # BillSrNo. : 112 or BillSrNo : 112
            r'Bill\s*Sr\s*No\.?\s*[:\s]*(\d+)',   # Bill Sr No. : 112
            r'Bill\s*Serial\s*No\.?\s*[:\s]*(\d+)' # Bill Serial No. : 112
        ], text)
    
    bill_data = {
        "bill_sr_no": bill_sr_no,
        "property_id": find_value([r'Property\s*Id[:\s]*([A-Z0-9]+)', r'PropertyId[:\s]*([A-Z0-9]+)'], text),
        "old_property_id": find_value([r'Old\s*Property\s*Id[:\s]*([A-Z0-9/-]+)', r'OldPropertyId[:\s]*([A-Z0-9/-]+)'], text),
        "financial_year": find_value([r'Financial\s*Year[:\s]*(\d{4}-\d{2,4})', r'FY[:\s]*(\d{4}-\d{2,4})'], text, "2025-26"),
        "print_date": find_value([r'Print\s*Date[:\s]*([0-9/\-]+)', r'Date[:\s]*([0-9/\-]+)'], text),
        "latitude": latitude,
        "longitude": longitude,
        "mobile": find_value([r'Mobile\s*No[:\s]*(\d{10})', r'Mobile[:\s]*(\d{10})', r'Phone[:\s]*(\d{10})'], text),
        "colony": find_value([r'Colony\s*Name[:\s]*([^\n]+)', r'Colony[:\s]*([^\n]+)'], text),
        "owner_name": find_value([r'Owner\s*Name[:\s]*([^\n]+)', r'Owner[:\s]*([^\n]+)'], text),
        "plot_address": find_value([r'Plot\s*Address[:\s]*([^\n]+)', r'Address[:\s]*([^\n]+)'], text),
        "permanent_address": find_value([r'Permanent\s*Address[:\s]*([^\n]+)'], text),
        "total_area": find_value([r'Total\s*Area[:\s]*([0-9.]+\s*SqYard)', r'Area[:\s]*([0-9.]+)'], text),
        "category": find_value([r'Category[:\s]*([^\n,]+)', r'Type[:\s]*([^\n,]+)'], text),
        "authorized_status": find_value([r'Authorized\s*Status[:\s]*([^\n]+)'], text),
        "total_outstanding": find_value([
            r'Total\s*Outstanding\s*as\s*on\s*date[^=]*=\s*([0-9,.-]+)',  # Total Outstanding as on date (PO+AO+...)= 3179.16
            r'Total\s*Outstanding[:\s]*Rs?\.?\s*([0-9,.-]+)', 
            r'Outstanding[:\s]*Rs?\.?\s*([0-9,.-]+)'
        ], text),
        "property_tax_outstanding": find_value([r'Property\s*&?\s*Fire\s*Tax\s*Outstanding[^\d]*([0-9,.-]+)'], text),
        "outstanding_property_arrear": find_value([r'Outstanding\s*Property\s*Tax\s*Arrear[^\d]*([0-9,.-]+)', r'AO[=:\s]*([0-9,.-]+)'], text),
        "outstanding_fire_arrear": find_value([r'Outstanding\s*Fire\s*Tax\s*Arrear[^\d]*([0-9,.-]+)', r'FO[=:\s]*([0-9,.-]+)'], text),
        "outstanding_interest": find_value([r'Outstanding\s*Interest\s*on\s*Arrear[^\d]*([0-9,.-]+)', r'IO[=:\s]*([0-9,.-]+)'], text),
        "outstanding_garbage": find_value([r'Outstanding\s*Garbage\s*Collection\s*Charges[^\d]*([0-9,.-]+)', r'SO1[=:\s]*([0-9,.-]+)'], text),
        "page_number": page_num
    }
    
    return bill_data

# Calculate distance between two GPS points (Haversine formula)
def haversine_distance(lat1, lon1, lat2, lon2):
    R = 6371000  # Earth radius in meters
    phi1 = math.radians(lat1)
    phi2 = math.radians(lat2)
    delta_phi = math.radians(lat2 - lat1)
    delta_lambda = math.radians(lon2 - lon1)
    a = math.sin(delta_phi/2)**2 + math.cos(phi1) * math.cos(phi2) * math.sin(delta_lambda/2)**2
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1-a))
    return R * c

# Sort bills by GPS route (nearest neighbor algorithm with location grouping)
def sort_by_gps_route(bills: list) -> list:
    if not bills:
        return bills
    
    # Separate bills with and without GPS
    valid_bills = [b for b in bills if b.get('latitude') and b.get('longitude')]
    no_gps_bills = [b for b in bills if not b.get('latitude') or not b.get('longitude')]
    
    if not valid_bills:
        return bills
    
    # Group bills by unique GPS location (round to 6 decimal places)
    location_groups = {}
    for bill in valid_bills:
        # Round coordinates to group nearby points (within ~0.1 meter)
        key = (round(bill['latitude'], 6), round(bill['longitude'], 6))
        if key not in location_groups:
            location_groups[key] = []
        location_groups[key].append(bill)
    
    # Get list of unique locations
    unique_locations = list(location_groups.keys())
    
    if len(unique_locations) <= 1:
        # All bills at same location, just return them
        return valid_bills + no_gps_bills
    
    # Find starting point - use the northwestern-most point (highest lat, lowest long)
    # This gives a consistent starting point
    start_idx = 0
    best_score = float('-inf')
    for i, loc in enumerate(unique_locations):
        score = loc[0] - loc[1] * 0.1  # Favor north and west
        if score > best_score:
            best_score = score
            start_idx = i
    
    # Sort unique locations using nearest neighbor algorithm
    sorted_locations = [unique_locations[start_idx]]
    remaining_locations = unique_locations[:start_idx] + unique_locations[start_idx+1:]
    
    while remaining_locations:
        last_loc = sorted_locations[-1]
        
        # Find nearest location
        nearest_idx = 0
        nearest_dist = float('inf')
        
        for i, loc in enumerate(remaining_locations):
            dist = haversine_distance(last_loc[0], last_loc[1], loc[0], loc[1])
            if dist < nearest_dist:
                nearest_dist = dist
                nearest_idx = i
        
        sorted_locations.append(remaining_locations.pop(nearest_idx))
    
    # Build final sorted list with all bills from each location in order
    sorted_bills = []
    for loc in sorted_locations:
        # Get all bills at this location and add them
        bills_at_loc = location_groups[loc]
        sorted_bills.extend(bills_at_loc)
    
    # Add bills without GPS at the end
    sorted_bills.extend(no_gps_bills)
    
    return sorted_bills

@api_router.post("/admin/bills/upload-pdf")
async def upload_pdf_bills(
    file: UploadFile = File(...),
    batch_name: str = Form(...),
    authorization: str = Form(...)
):
    """Upload multi-page PDF and extract bill data from each page (ADMIN only)"""
    current_user = await get_current_user(authorization)
    if current_user["role"] != "ADMIN":
        raise HTTPException(status_code=403, detail="Only Admin can upload PDF bills")
    
    if not file.filename.lower().endswith('.pdf'):
        raise HTTPException(status_code=400, detail="Please upload a PDF file")
    
    # Save uploaded PDF
    content = await file.read()
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    pdf_filename = f"bills_{timestamp}.pdf"
    pdf_path = UPLOAD_DIR / pdf_filename
    
    async with aiofiles.open(pdf_path, 'wb') as f:
        await f.write(content)
    
    # Create batch record
    batch_id = str(uuid.uuid4())
    batch_doc = {
        "id": batch_id,
        "name": batch_name,
        "type": "PDF_BILLS",
        "pdf_filename": pdf_filename,
        "pdf_url": f"/api/uploads/{pdf_filename}",
        "uploaded_by": current_user["id"],
        "uploaded_at": datetime.now(timezone.utc).isoformat(),
        "status": "ACTIVE",
        "total_records": 0,
        "skip_stats": {
            "skipped_na_empty": 0,
            "skipped_vacant": 0,
            "na_serial_count": 0
        }
    }
    
    # Helper function to check if owner name is valid (not NA or empty)
    def is_valid_owner_name(name):
        """Check if owner name is valid - be LENIENT to avoid skipping good data"""
        if not name:
            return False
        name_clean = name.strip()
        name_upper = name_clean.upper()
        # Only skip if EXACTLY matches these invalid values
        invalid_values = ['NA', 'N/A', 'N.A.', '-', '--', '']
        return name_upper not in invalid_values and len(name_clean) > 0
    
    def should_skip_record(owner_name, category=""):
        """Determine if a record should be skipped - be CONSERVATIVE"""
        owner = (owner_name or "").strip().lower()
        cat = (category or "").strip().lower()
        
        # Only skip if owner is completely empty/invalid AND category is vacant
        if not owner or owner in ['na', 'n/a', '-', '--']:
            # Even if owner is NA, keep residential properties
            if 'residential' in cat or 'commercial' in cat:
                return False
            return True
        
        # Skip ONLY if explicitly marked as vacant plot with no owner
        if ('vacant' in cat or 'empty' in cat) and (not owner or len(owner) <= 2):
            return True
        
        return False
    
    # Extract text from each page using PyMuPDF
    bills = []
    skipped_count = 0
    na_serial_count = 0
    
    # Load self-certified PIDs from database for matching
    self_certified_docs = await db.self_certified_pids.find({}, {"pid": 1, "_id": 0}).to_list(None)
    self_certified_pids = set(doc["pid"].upper() for doc in self_certified_docs if doc.get("pid"))
    self_certified_count = 0
    not_self_certified_count = 0
    
    try:
        pdf_doc = fitz.open(str(pdf_path))
        
        # Vacant plot keywords to check - be more specific
        VACANT_KEYWORDS = ["vacant plot", "empty plot", "खाली प्लॉट"]
        
        def is_vacant_plot(owner_name, category=""):
            """Check if a bill represents a vacant plot - be STRICT, only skip obvious vacant plots"""
            owner = (owner_name or "").strip().lower()
            cat = (category or "").strip().lower()
            
            # Only skip if category explicitly says vacant AND owner is empty/invalid
            if "vacant" in cat and (not owner or owner in ['na', 'n/a', '-', 'nil']):
                return True
            
            # Check for exact vacant plot phrases
            for keyword in VACANT_KEYWORDS:
                if keyword in owner:
                    return True
            
            return False
        
        skipped_vacant = 0
        
        # First pass: Extract all bill data - BE LENIENT, include most records
        for page_num in range(len(pdf_doc)):
            page = pdf_doc[page_num]
            text = page.get_text()
            
            # Extract bill data (pass page for block-based BillSrNo extraction)
            bill_data = extract_bill_data(text, page_num + 1, page)
            
            owner_name = bill_data.get("owner_name", "")
            category = bill_data.get("category", "")
            
            # Use conservative skip logic - only skip obvious invalid records
            if should_skip_record(owner_name, category):
                skipped_count += 1
                continue
            
            # Skip only obvious vacant plots with no owner
            if is_vacant_plot(owner_name, category):
                skipped_vacant += 1
                continue
            
            bill_data["id"] = str(uuid.uuid4())
            bill_data["batch_id"] = batch_id
            bill_data["page_num"] = page_num + 1  # Store original page number
            
            # Check if serial number is valid
            pdf_serial = bill_data.get("bill_sr_no", "").strip()
            if pdf_serial and pdf_serial.isdigit():
                bill_data["serial_number"] = int(pdf_serial)
                bill_data["serial_na"] = False
                bill_data["bill_sr_no"] = pdf_serial
            else:
                bill_data["serial_number"] = 0
                bill_data["serial_na"] = True
                bill_data["bill_sr_no"] = "N0"  # Temporary, will be updated in second pass
                na_serial_count += 1
            
            bill_data["created_at"] = datetime.now(timezone.utc).isoformat()
            bill_data["status"] = "Pending"
            bill_data["gps_arranged"] = False
            
            # Check if this property is self-certified
            bill_prop_id = bill_data.get("property_id", "")
            is_self_certified = bill_prop_id.upper() in self_certified_pids if bill_prop_id else False
            bill_data["self_certified"] = is_self_certified
            if is_self_certified:
                self_certified_count += 1
            else:
                not_self_certified_count += 1
            
            bills.append(bill_data)
        
        pdf_doc.close()
        
        # Second pass: Fix N/A serials to use nearest valid serial BY GPS LOCATION
        # Get all valid serial numbers with their GPS coordinates
        valid_serials_with_gps = []
        for i, b in enumerate(bills):
            if not b["serial_na"] and b.get("latitude") and b.get("longitude"):
                valid_serials_with_gps.append({
                    "serial": b["serial_number"],
                    "lat": b["latitude"],
                    "lng": b["longitude"]
                })
        
        if valid_serials_with_gps:
            for i, bill in enumerate(bills):
                if bill["serial_na"] and bill.get("latitude") and bill.get("longitude"):
                    # Find the nearest valid serial based on GPS distance
                    nearest_serial = 0
                    min_distance = float('inf')
                    
                    bill_lat = bill["latitude"]
                    bill_lng = bill["longitude"]
                    
                    for vs in valid_serials_with_gps:
                        # Calculate simple Euclidean distance (good enough for nearby points)
                        dist = ((vs["lat"] - bill_lat) ** 2 + (vs["lng"] - bill_lng) ** 2) ** 0.5
                        if dist < min_distance:
                            min_distance = dist
                            nearest_serial = vs["serial"]
                    
                    bill["bill_sr_no"] = f"N{nearest_serial}"
                elif bill["serial_na"]:
                    # No GPS, use first valid serial as fallback
                    bill["bill_sr_no"] = f"N{valid_serials_with_gps[0]['serial'] if valid_serials_with_gps else 0}"
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error processing PDF: {str(e)}")
    
    # Build upload message
    upload_message = f"Uploaded {len(bills)} bills. Skipped {skipped_count} records with NA/empty owner names."
    if skipped_vacant > 0:
        upload_message += f" Skipped {skipped_vacant} vacant plots."
    if self_certified_count > 0:
        upload_message += f" {self_certified_count} self-certified."
    if not_self_certified_count > 0:
        upload_message += f" {not_self_certified_count} not self-certified."
    
    # Insert bills into database
    if bills:
        await db.bills.insert_many(bills)
        batch_doc["total_records"] = len(bills)
        batch_doc["skip_stats"] = {
            "skipped_na_empty": skipped_count,
            "skipped_vacant": skipped_vacant,
            "na_serial_count": na_serial_count,
            "total_skipped": skipped_count + skipped_vacant,
            "self_certified_count": self_certified_count,
            "not_self_certified_count": not_self_certified_count,
            "upload_message": upload_message
        }
    
    await db.batches.insert_one(batch_doc)
    
    # Get unique colonies with their bill counts
    colony_stats = {}
    for b in bills:
        colony_name = b.get("colony", "").strip()
        if colony_name:
            if colony_name not in colony_stats:
                colony_stats[colony_name] = {"total": 0, "na_serial": 0}
            colony_stats[colony_name]["total"] += 1
            if b.get("serial_na"):
                colony_stats[colony_name]["na_serial"] += 1
    
    # Get unique colonies
    colonies = list(set([b.get("colony", "").strip() for b in bills if b.get("colony")]))
    
    return {
        "batch_id": batch_id,
        "name": batch_name,
        "total_bills": len(bills),
        "skipped_bills": skipped_count,
        "skipped_vacant": skipped_vacant,
        "na_serial_bills": na_serial_count,
        "self_certified": self_certified_count,
        "not_self_certified": not_self_certified_count,
        "colonies": colonies,
        "message": upload_message
    }

@api_router.get("/admin/bills/export-excel")
async def export_bills_excel(
    batch_id: Optional[str] = None,
    colony: Optional[str] = None,
    self_cert_filter: Optional[str] = None,  # 'self_certified', 'not_self_certified', 'all'
    current_user: dict = Depends(get_current_user)
):
    """Export bills to Excel with optional self-certification filter"""
    if current_user["role"] not in ADMIN_VIEW_ROLES:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Build query
    query = {}
    if batch_id and batch_id.strip():
        query["batch_id"] = batch_id.strip()
    if colony and colony.strip():
        query["colony"] = colony.strip()
    
    # Apply self-certification filter
    if self_cert_filter == "self_certified":
        query["self_certified"] = True
    elif self_cert_filter == "not_self_certified":
        query["$or"] = [{"self_certified": False}, {"self_certified": {"$exists": False}}]
    
    # Get bills
    bills = await db.bills.find(query, {"_id": 0}).to_list(None)
    
    # Get count for the filter description
    filter_desc = "All Bills"
    if self_cert_filter == "self_certified":
        filter_desc = "Self-Certified Only"
    elif self_cert_filter == "not_self_certified":
        filter_desc = "Not Self-Certified Only"
    
    if not bills:
        raise HTTPException(
            status_code=404, 
            detail=f"No bills found for '{filter_desc}' filter. Colony: {colony or 'All'}. Please check if there are any {filter_desc.lower()} records."
        )
    
    # Create DataFrame
    df_data = []
    for bill in bills:
        df_data.append({
            "Serial No": bill.get("serial_number", ""),
            "Bill Sr No": bill.get("bill_sr_no", ""),
            "Property ID": bill.get("property_id", ""),
            "Owner Name": bill.get("owner_name", ""),
            "Mobile": bill.get("mobile", ""),
            "Colony": bill.get("colony", ""),
            "Category": bill.get("category", ""),
            "Plot Address": bill.get("plot_address", ""),
            "Total Area": bill.get("total_area", ""),
            "Total Outstanding": bill.get("total_outstanding", ""),
            "Self Certified": "Yes" if bill.get("self_certified") else "No",
            "Serial NA": "Yes" if bill.get("serial_na") else "No",
            "Latitude": bill.get("latitude", ""),
            "Longitude": bill.get("longitude", "")
        })
    
    df = pd.DataFrame(df_data)
    
    # Generate filename
    filter_suffix = ""
    if self_cert_filter == "self_certified":
        filter_suffix = "_self_certified"
    elif self_cert_filter == "not_self_certified":
        filter_suffix = "_not_self_certified"
    
    colony_suffix = f"_{colony.replace(' ', '_')}" if colony and colony.strip() else ""
    filename = f"bills_export{colony_suffix}{filter_suffix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    export_path = UPLOAD_DIR / filename
    
    df.to_excel(str(export_path), index=False)
    
    return FileResponse(
        path=str(export_path),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=filename
    )

@api_router.get("/admin/bills")
async def list_bills(
    batch_id: Optional[str] = None,
    colony: Optional[str] = None,
    status: Optional[str] = None,
    sorted_by_route: bool = False,
    page: int = 1,
    limit: int = 50,
    current_user: dict = Depends(get_current_user)
):
    """Get bills with optional filtering"""
    if current_user["role"] not in ADMIN_VIEW_ROLES:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    query = {}
    if batch_id and batch_id.strip():
        query["batch_id"] = batch_id
    if colony and colony.strip():
        query["colony"] = {"$regex": colony, "$options": "i"}
    if status and status.strip():
        query["status"] = status
    
    total = await db.bills.count_documents(query)
    
    if sorted_by_route:
        # Get all matching bills and sort by GPS route
        all_bills = await db.bills.find(query, {"_id": 0}).to_list(None)
        sorted_bills = sort_by_gps_route(all_bills)
        
        # Assign new serial numbers
        for i, bill in enumerate(sorted_bills):
            bill["route_serial"] = i + 1
        
        # Paginate
        start = (page - 1) * limit
        bills = sorted_bills[start:start + limit]
    else:
        # Sort by page_number to maintain original PDF sequence
        bills = await db.bills.find(query, {"_id": 0}).sort("page_number", 1).skip((page - 1) * limit).limit(limit).to_list(limit)
    
    return {
        "bills": bills,
        "total": total,
        "page": page,
        "pages": (total + limit - 1) // limit
    }

@api_router.get("/admin/bills/colonies")
async def get_bill_colonies(
    batch_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get unique colonies from bills"""
    if current_user["role"] not in ADMIN_VIEW_ROLES:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    query = {}
    if batch_id and batch_id.strip():
        query["batch_id"] = batch_id
    
    colonies = await db.bills.distinct("colony", query)
    colonies = [c for c in colonies if c and c.strip()]
    
    return {"colonies": sorted(colonies)}

@api_router.get("/admin/bills/batch-stats/{batch_id}")
async def get_batch_stats(
    batch_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Get batch statistics including skip information"""
    if current_user["role"] not in ADMIN_VIEW_ROLES:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    batch = await db.batches.find_one({"id": batch_id}, {"_id": 0})
    if not batch:
        raise HTTPException(status_code=404, detail="Batch not found")
    
    # Get colony-wise stats
    pipeline = [
        {"$match": {"batch_id": batch_id}},
        {"$group": {
            "_id": "$colony",
            "total": {"$sum": 1},
            "na_serial": {"$sum": {"$cond": ["$serial_na", 1, 0]}},
            "valid_serial": {"$sum": {"$cond": ["$serial_na", 0, 1]}}
        }},
        {"$sort": {"_id": 1}}
    ]
    colony_stats = await db.bills.aggregate(pipeline).to_list(None)
    
    return {
        "batch": batch,
        "skip_stats": batch.get("skip_stats", {}),
        "colony_stats": [
            {
                "colony": stat["_id"] or "Unknown",
                "total_bills": stat["total"],
                "na_serial_bills": stat["na_serial"],
                "valid_serial_bills": stat["valid_serial"]
            }
            for stat in colony_stats
        ]
    }

@api_router.get("/admin/bills/colony-stats/{colony_name}")
async def get_colony_stats(
    colony_name: str,
    current_user: dict = Depends(get_current_user)
):
    """Get statistics for a specific colony including skip information"""
    if current_user["role"] not in ADMIN_VIEW_ROLES:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    from urllib.parse import unquote
    colony_name = unquote(colony_name)
    
    # Get total bills in colony
    total_bills = await db.bills.count_documents({"colony": colony_name})
    
    # Get NA serial count
    na_serial_count = await db.bills.count_documents({"colony": colony_name, "serial_na": True})
    
    # Get valid serial count
    valid_serial_count = await db.bills.count_documents({"colony": colony_name, "serial_na": {"$ne": True}, "serial_number": {"$gt": 0}})
    
    # Get bills with GPS
    with_gps = await db.bills.count_documents({"colony": colony_name, "latitude": {"$exists": True, "$ne": None}})
    
    # Get self-certified counts
    self_certified_count = await db.bills.count_documents({"colony": colony_name, "self_certified": True})
    not_self_certified_count = await db.bills.count_documents({
        "colony": colony_name, 
        "$or": [{"self_certified": False}, {"self_certified": {"$exists": False}}]
    })
    
    # Get category breakdown
    category_pipeline = [
        {"$match": {"colony": colony_name}},
        {"$group": {"_id": "$category", "count": {"$sum": 1}}},
        {"$sort": {"count": -1}}
    ]
    category_stats = await db.bills.aggregate(category_pipeline).to_list(None)
    
    # Get batch info for this colony (to get skip stats and messages)
    batch_ids = await db.bills.distinct("batch_id", {"colony": colony_name})
    skip_stats = {"skipped_na_empty": 0, "skipped_vacant": 0, "na_serial_count": na_serial_count}
    upload_messages = []
    add_to_properties_messages = []
    
    for batch_id in batch_ids:
        if batch_id:
            batch = await db.batches.find_one(
                {"id": batch_id}, 
                {"_id": 0, "name": 1, "skip_stats": 1, "add_to_properties_stats": 1}
            )
            if batch:
                batch_name = batch.get("name", "Unknown Batch")
                # Get PDF upload stats
                if batch.get("skip_stats"):
                    bs = batch["skip_stats"]
                    skip_stats["skipped_na_empty"] += bs.get("skipped_na_empty", 0)
                    skip_stats["skipped_vacant"] += bs.get("skipped_vacant", 0)
                    if bs.get("upload_message"):
                        upload_messages.append({
                            "batch_name": batch_name,
                            "message": bs["upload_message"]
                        })
                # Get Add to Properties stats
                if batch.get("add_to_properties_stats"):
                    aps = batch["add_to_properties_stats"]
                    if aps.get("message"):
                        add_to_properties_messages.append({
                            "batch_name": batch_name,
                            "message": aps["message"],
                            "stats": aps
                        })
    
    return {
        "colony": colony_name,
        "total_bills": total_bills,
        "na_serial_count": na_serial_count,
        "valid_serial_count": valid_serial_count,
        "with_gps": with_gps,
        "self_certified_count": self_certified_count,
        "not_self_certified_count": not_self_certified_count,
        "skip_stats": skip_stats,
        "upload_messages": upload_messages,
        "add_to_properties_messages": add_to_properties_messages,
        "category_breakdown": [
            {"category": stat["_id"] or "Unknown", "count": stat["count"]}
            for stat in category_stats
        ]
    }

@api_router.put("/admin/bills/{bill_id}")
async def update_bill(
    bill_id: str,
    current_user: dict = Depends(get_current_user),
    owner_name: str = Form(None),
    mobile: str = Form(None),
    plot_address: str = Form(None),
    permanent_address: str = Form(None),
    category: str = Form(None),
    total_area: str = Form(None),
    total_outstanding: str = Form(None),
    colony: str = Form(None)
):
    """Edit bill data"""
    if current_user["role"] not in ADMIN_ROLES:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    bill = await db.bills.find_one({"id": bill_id})
    if not bill:
        raise HTTPException(status_code=404, detail="Bill not found")
    
    update_data = {"updated_at": datetime.now(timezone.utc).isoformat()}
    if owner_name is not None:
        update_data["owner_name"] = owner_name
    if mobile is not None:
        update_data["mobile"] = mobile
    if plot_address is not None:
        update_data["plot_address"] = plot_address
    if permanent_address is not None:
        update_data["permanent_address"] = permanent_address
    if category is not None:
        update_data["category"] = category
    if total_area is not None:
        update_data["total_area"] = total_area
    if total_outstanding is not None:
        update_data["total_outstanding"] = total_outstanding
    if colony is not None:
        update_data["colony"] = colony
    
    await db.bills.update_one({"id": bill_id}, {"$set": update_data})
    
    return {"message": "Bill updated successfully"}

@api_router.post("/admin/bills/arrange-by-route")
async def arrange_bills_by_route(
    batch_id: str = Form(None),
    colony: str = Form(None),
    current_user: dict = Depends(get_current_user)
):
    """Arrange bills by GPS route and assign new serial numbers (excludes N/A serial bills)"""
    if current_user["role"] not in ADMIN_ROLES:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    query = {}
    if batch_id and batch_id.strip():
        query["batch_id"] = batch_id
    if colony and colony.strip():
        query["colony"] = {"$regex": colony, "$options": "i"}
    
    # Get all matching bills
    all_bills = await db.bills.find(query, {"_id": 0}).to_list(None)
    
    if not all_bills:
        raise HTTPException(status_code=404, detail="No bills found")
    
    # Separate bills with valid serial numbers from N/A serial bills
    valid_bills = [b for b in all_bills if not b.get("serial_na", False)]
    na_bills = [b for b in all_bills if b.get("serial_na", False)]
    
    if not valid_bills:
        raise HTTPException(status_code=400, detail="No bills with valid serial numbers to arrange")
    
    # Sort valid bills by GPS route
    sorted_bills = sort_by_gps_route(valid_bills)
    
    # Update serial numbers for sorted bills and mark as GPS arranged
    for i, bill in enumerate(sorted_bills):
        await db.bills.update_one(
            {"id": bill["id"]},
            {"$set": {"serial_number": i + 1, "gps_arranged": True}}
        )
    
    # Mark N/A bills as skipped from ordering
    for bill in na_bills:
        await db.bills.update_one(
            {"id": bill["id"]},
            {"$set": {"gps_arranged": False, "skipped_from_order": True}}
        )
    
    return {
        "message": f"Arranged {len(sorted_bills)} bills by GPS route. {len(na_bills)} bills with N/A serial skipped.",
        "total_arranged": len(sorted_bills),
        "skipped_na": len(na_bills)
    }

@api_router.post("/admin/bills/generate-pdf")
async def generate_arranged_pdf(
    batch_id: str = Form(None),
    colony: str = Form(None),
    bills_per_page: int = Form(1),  # 1 = full page, 3 = stacked vertically
    print_serial: str = Form("true"),  # Print serial number on PDF
    self_certified_filter: str = Form("all"),  # "all", "self_certified", "not_self_certified"
    current_user: dict = Depends(get_current_user)
):
    """Generate PDF with invoices. 3 per page = landscape bills scaled & stacked vertically on A4."""
    if current_user["role"] != "ADMIN":
        raise HTTPException(status_code=403, detail="Only Admin can generate PDF")
    
    should_print_serial = print_serial.lower() == "true"
    
    query = {}
    if batch_id and batch_id.strip():
        query["batch_id"] = batch_id
    if colony and colony.strip():
        query["colony"] = {"$regex": colony, "$options": "i"}
    
    # Add self_certified filter
    if self_certified_filter == "self_certified":
        query["self_certified"] = True
    elif self_certified_filter == "not_self_certified":
        query["self_certified"] = {"$ne": True}
    
    bills = await db.bills.find(query, {"_id": 0}).sort("page_number", 1).to_list(None)
    
    if not bills:
        raise HTTPException(status_code=404, detail="No bills found")
    
    # Helper functions to skip vacant plots and invalid owner names (same as Add to Properties)
    def should_skip_for_pdf(bill):
        """Skip vacant plots and bills with no valid owner name"""
        owner = (bill.get("owner_name") or "").strip().lower()
        category = (bill.get("category") or "").strip().lower()
        
        # Skip if no owner name or invalid owner
        if not owner or owner in ['na', 'n/a', 'n.a.', '-', '--', 'nil', 'none']:
            return True
        
        # Skip vacant plots
        if "vacant" in category or "empty" in category:
            return True
        if "vacant" in owner or "empty plot" in owner or "खाली" in owner:
            return True
        
        return False
    
    # Filter out vacant plots and invalid owner names
    valid_bills = [b for b in bills if not should_skip_for_pdf(b)]
    skipped_count = len(bills) - len(valid_bills)
    
    if not valid_bills:
        raise HTTPException(status_code=404, detail=f"No valid bills found (skipped {skipped_count} vacant/invalid records)")
    
    # Keep original bills for serial number lookup (N/A serials need ALL valid serials for nearby lookup)
    all_bills_for_serial_lookup = bills
    bills = valid_bills  # Use filtered bills for PDF generation
    
    batch = await db.batches.find_one({"id": bills[0]["batch_id"]})
    if not batch or not batch.get("pdf_filename"):
        raise HTTPException(status_code=404, detail="Original PDF not found")
    
    original_pdf_path = UPLOAD_DIR / batch["pdf_filename"]
    if not original_pdf_path.exists():
        raise HTTPException(status_code=404, detail="Original PDF file not found")
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_filename = f"arranged_{colony or 'all'}_{timestamp}.pdf"
    output_path = UPLOAD_DIR / output_filename
    
    src_pdf = fitz.open(str(original_pdf_path))
    output_pdf = fitz.open()
    
    # Build serial number lookup from ALL bills (including skipped ones)
    # This ensures N/A serials can find nearby valid serials
    valid_serials_with_gps = []
    for b in all_bills_for_serial_lookup:
        if not b.get("serial_na", False) and b.get("serial_number", 0) > 0 and b.get("latitude") and b.get("longitude"):
            valid_serials_with_gps.append({
                "serial": b["serial_number"],
                "lat": b["latitude"],
                "lng": b["longitude"]
            })
    
    def get_display_serial(bill):
        """Get display serial number:
        - If bill has a valid serial_number (not 0, not NA) → use that number (e.g., 7, 42)
        - If serial is NA/blank/0 → find nearest property with valid serial based on GPS and prefix with N (e.g., N7)
        """
        bill_serial = bill.get("serial_number") or 0
        is_serial_na = bill.get("serial_na", False) or bill_serial == 0 or bill_serial is None
        
        if not is_serial_na and bill_serial > 0:
            # Has valid serial number - use it directly (e.g., 7, 42, 156)
            return str(int(bill_serial))
        else:
            # Serial is NA/blank - find nearest property based on GPS and prefix with N
            nearest_serial = 0
            if valid_serials_with_gps and bill.get("latitude") and bill.get("longitude"):
                min_distance = float('inf')
                bill_lat = float(bill["latitude"])
                bill_lng = float(bill["longitude"])
                
                for vs in valid_serials_with_gps:
                    dist = ((vs["lat"] - bill_lat) ** 2 + (vs["lng"] - bill_lng) ** 2) ** 0.5
                    if dist < min_distance:
                        min_distance = dist
                        nearest_serial = vs["serial"]
            elif valid_serials_with_gps:
                # Fallback to first valid serial if no GPS on this bill
                nearest_serial = valid_serials_with_gps[0]["serial"]
            
            # Return N-prefix with nearest serial (e.g., N7, N42)
            if nearest_serial > 0:
                return f"N{nearest_serial}"
            else:
                return "N/A"
    
    included_count = 0
    
    if bills_per_page == 1:
        # ONE BILL PER PAGE - Copy original page directly, add serial overlay
        for bill in bills:
            page_num = bill.get("page_number", 1) - 1
            if page_num < 0 or page_num >= len(src_pdf):
                continue
            
            # Get source page
            src_page = src_pdf[page_num]
            
            # COPY page directly to preserve original quality and size
            output_pdf.insert_pdf(src_pdf, from_page=page_num, to_page=page_num)
            new_page = output_pdf[-1]  # Get the newly inserted page
            
            # Get page rotation and dimensions
            rotation = new_page.rotation
            rect = new_page.rect
            
            # Add Hindi message FIRST (left side), then serial number (right side)
            # Both at top with 50px padding
            is_self_certified = bill.get("self_certified", False)
            
            # Load font for Hindi + English support
            # Try Gargi font for proper Hindi rendering
            gargi_font = '/usr/share/fonts/truetype/Gargi/Gargi.ttf'
            samyak_font = '/usr/share/fonts/truetype/samyak-fonts/Samyak-Devanagari.ttf'
            freesans_font = '/usr/share/fonts/truetype/freefont/FreeSans.ttf'
            
            if os.path.exists(gargi_font):
                new_page.insert_font(fontname='gargi', fontbuffer=open(gargi_font, 'rb').read())
                font_name = 'gargi'
            elif os.path.exists(samyak_font):
                new_page.insert_font(fontname='samyak', fontbuffer=open(samyak_font, 'rb').read())
                font_name = 'samyak'
            elif os.path.exists(freesans_font):
                new_page.insert_font(fontname='freesans', fontbuffer=open(freesans_font, 'rb').read())
                font_name = 'freesans'
            else:
                font_name = 'helv'
            
            # Add serial number (LEFT side)
            if should_print_serial:
                serial_text = get_display_serial(bill)
                font_size = 18
                
                # Position: LEFT side at top with 50px padding
                if rotation == 90:
                    visual_point = fitz.Point(80, 50)  # Left side, 50px from top
                    internal_point = visual_point * new_page.derotation_matrix
                    text_rotate = 90
                elif rotation == 270:
                    visual_point = fitz.Point(rect.width - 80, rect.height - 50)
                    internal_point = visual_point * new_page.derotation_matrix
                    text_rotate = 270
                else:
                    internal_point = fitz.Point(80, 50)
                    text_rotate = 0
                
                # Insert serial number - RED BOLD
                new_page.insert_text(
                    internal_point,
                    serial_text,
                    fontsize=font_size,
                    fontname="helv",
                    color=(1, 0, 0),
                    rotate=text_rotate
                )
            
            # Add note for non-self-certified properties - using pre-generated image
            if not is_self_certified:
                # Use pre-generated Hindi note image
                note_img_path = "/tmp/hindi_note_cached.png"
                
                # Generate image only if it doesn't exist
                if not os.path.exists(note_img_path):
                    import subprocess
                    import tempfile
                    
                    hindi_note_html = '''<!DOCTYPE html>
<html><head><meta charset="UTF-8">
<style>body{margin:0;padding:2px 5px;font-family:'Noto Sans Devanagari','Lohit Devanagari',sans-serif;font-size:22px;font-weight:bold;color:#cc0000;background:transparent;white-space:nowrap;}</style>
</head><body>Note : आप अपनी प्रॉपर्टी ID को सेल्फ सर्टिफाइड करवाए, जिससे कि आपकी प्रॉपर्टी ID के साथ कोई छेड़ -छाड़ ना कर सके।</body></html>'''
                    
                    with tempfile.NamedTemporaryFile(mode='w', suffix='.html', delete=False, encoding='utf-8') as f:
                        f.write(hindi_note_html)
                        html_path = f.name
                    
                    try:
                        subprocess.run([
                            'xvfb-run', '--auto-servernum', 'wkhtmltoimage',
                            '--encoding', 'utf-8', '--width', '1150', '--height', '50', '--quality', '100',
                            html_path, note_img_path
                        ], capture_output=True, timeout=30)
                    finally:
                        if os.path.exists(html_path):
                            os.unlink(html_path)
                
                if os.path.exists(note_img_path):
                    try:
                        # Insert image into PDF - TOP LEFT position
                        # Page is landscape (rotated 90), so adjust coordinates
                        if rotation == 90:
                            # Rotated page - place at TOP LEFT (visually)
                            img_rect = fitz.Rect(10, 10, 60, 500)
                        elif rotation == 270:
                            img_rect = fitz.Rect(rect.width - 60, 10, rect.width - 10, 500)
                        else:
                            # Normal page - TOP LEFT position
                            img_rect = fitz.Rect(10, 10, 500, 60)
                        
                        new_page.insert_image(img_rect, filename=note_img_path, rotate=rotation)
                    except Exception as e:
                        logger.error(f"Error inserting Hindi note image: {e}")
            
            included_count += 1
    else:
        # 2 OR 3 BILLS PER PAGE - Compact & Print quality optimized
        # A4 dimensions - slightly reduced for compact output
        A4_WIDTH = 560  # Reduced from 595.28
        A4_HEIGHT = 792  # Reduced from 841.89
        
        # Use the requested bills_per_page (2 or 3)
        num_bills = bills_per_page if bills_per_page in [2, 3] else 3
        
        # Each slot height
        slot_height = A4_HEIGHT / num_bills
        
        current_page = None
        position = 0
        
        for bill in bills:
            page_num = bill.get("page_number", 1) - 1
            if page_num < 0 or page_num >= len(src_pdf):
                continue
            
            if position == 0:
                current_page = output_pdf.new_page(width=A4_WIDTH, height=A4_HEIGHT)
            
            # Get source page and render to image
            src_page = src_pdf[page_num]
            
            # Render at 1.5x scale for good quality (150 DPI) - balanced quality/size
            mat = fitz.Matrix(1.5, 1.5)
            pix = src_page.get_pixmap(matrix=mat, alpha=False)
            
            # Use JPEG format with 85% quality - much smaller than PNG, good quality
            img_bytes = pix.tobytes("jpeg", jpg_quality=85)
            
            # Pixmap dimensions
            pix_width = pix.width
            pix_height = pix.height
            
            # COMPACT - efficient use of space, reduced margins
            if num_bills == 2:
                # 2 bills per page - compact
                available_width = A4_WIDTH - 4
                available_height = slot_height - 2
                scale_boost = 0.88
            else:
                # 3 bills per page - very compact
                available_width = A4_WIDTH - 4
                available_height = slot_height - 1
                scale_boost = 1.30
            
            scale_w = available_width / pix_width
            scale_h = available_height / pix_height
            scale = min(scale_w, scale_h) * scale_boost
            
            final_width = pix_width * scale
            final_height = pix_height * scale
            
            # Center in slot - minimal vertical gap
            x_offset = (A4_WIDTH - final_width) / 2
            y_start = position * slot_height
            y_offset = (slot_height - final_height) / 2 * 0.3
            
            rect = fitz.Rect(
                x_offset,
                y_start + y_offset,
                x_offset + final_width,
                y_start + y_offset + final_height
            )
            
            # Insert the JPEG image
            current_page.insert_image(rect, stream=img_bytes)
            
            # Add serial number overlay if enabled
            if should_print_serial:
                serial_text = get_display_serial(bill)
                
                # Draw serial number - compact and readable
                font_size = 10 if num_bills == 3 else 11
                text_x = rect.x1 - len(serial_text) * font_size * 0.5 - 3
                text_y = rect.y0 + font_size + 1
                
                # White background rectangle
                bg_rect = fitz.Rect(text_x - 2, rect.y0 + 1, rect.x1 - 1, text_y + 1)
                current_page.draw_rect(bg_rect, color=(1, 1, 1), fill=(1, 1, 1))
                
                # Red bold text
                current_page.insert_text(
                    (text_x, text_y),
                    serial_text,
                    fontsize=font_size,
                    fontname="hebo",  # Helvetica Bold
                    color=(1, 0, 0)
                )
            
            included_count += 1
            position = (position + 1) % num_bills
    
    # SAVE WITH COMPRESSION - smaller file size
    output_pdf.save(
        str(output_path),
        garbage=4,           # Maximum garbage collection
        deflate=True,        # Compress streams
        clean=True,          # Clean unused objects
        deflate_images=True, # Compress images
        deflate_fonts=True   # Compress fonts
    )
    output_pdf.close()
    src_pdf.close()
    
    pages_created = (included_count + bills_per_page - 1) // bills_per_page if bills_per_page > 1 else included_count
    
    return {
        "message": f"Generated {pages_created} pages with {included_count} bills ({bills_per_page} per page)",
        "filename": output_filename,
        "download_url": f"/api/uploads/{output_filename}"
    }

@api_router.post("/admin/bills/split-by-employee")
async def split_bills_by_employee(
    batch_id: str = Form(None),
    colony: str = Form(None),
    employee_count: int = Form(...),
    sn_font_size: int = Form(48),
    sn_color: str = Form("red"),
    current_user: dict = Depends(get_current_user)
):
    """Split bills into separate PDFs for each employee"""
    if current_user["role"] not in ADMIN_ROLES:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    if employee_count < 1 or employee_count > 100:
        raise HTTPException(status_code=400, detail="Employee count must be between 1 and 100")
    
    query = {}
    if batch_id and batch_id.strip():
        query["batch_id"] = batch_id
    if colony and colony.strip():
        query["colony"] = {"$regex": colony, "$options": "i"}
    
    # Get arranged bills
    all_bills = await db.bills.find(query, {"_id": 0}).sort("serial_number", 1).to_list(None)
    
    if not all_bills:
        raise HTTPException(status_code=404, detail="No bills found")
    
    # Skip vacant plots and invalid owner names (same logic as generate-pdf)
    def should_skip_for_pdf(bill):
        owner = (bill.get("owner_name") or "").strip().lower()
        category = (bill.get("category") or "").strip().lower()
        if not owner or owner in ['na', 'n/a', 'n.a.', '-', '--', 'nil', 'none']:
            return True
        if "vacant" in category or "empty" in category:
            return True
        if "vacant" in owner or "empty plot" in owner or "खाली" in owner:
            return True
        return False
    
    bills = [b for b in all_bills if not should_skip_for_pdf(b)]
    
    if not bills:
        raise HTTPException(status_code=404, detail="No valid bills found after filtering")
    
    # Get original PDF
    batch = await db.batches.find_one({"id": bills[0]["batch_id"]})
    if not batch or not batch.get("pdf_filename"):
        raise HTTPException(status_code=404, detail="Original PDF not found")
    
    original_pdf_path = UPLOAD_DIR / batch["pdf_filename"]
    if not original_pdf_path.exists():
        raise HTTPException(status_code=404, detail="Original PDF file not found")
    
    # Calculate bills per employee
    total_bills = len(bills)
    bills_per_employee = math.ceil(total_bills / employee_count)
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    generated_files = []
    
    # Color mapping
    color_map = {
        "red": (1, 0, 0),
        "blue": (0, 0, 1),
        "green": (0, 0.5, 0),
        "black": (0, 0, 0),
        "orange": (1, 0.5, 0)
    }
    sn_rgb = color_map.get(sn_color.lower(), (1, 0, 0))
    
    src_pdf = fitz.open(str(original_pdf_path))
    
    # Build serial number lookup from ALL bills for N/A serials
    valid_serials_with_gps = []
    for b in all_bills:
        if not b.get("serial_na", False) and b.get("serial_number", 0) > 0 and b.get("latitude") and b.get("longitude"):
            valid_serials_with_gps.append({
                "serial": b["serial_number"],
                "lat": b["latitude"],
                "lng": b["longitude"]
            })
    
    def get_display_serial(bill):
        """Get display serial number:
        - If bill has a valid serial_number (not 0, not NA) → use that number (e.g., 7, 42)
        - If serial is NA/blank/0 → find nearest property with valid serial based on GPS and prefix with N (e.g., N7)
        """
        bill_serial = bill.get("serial_number") or 0
        is_serial_na = bill.get("serial_na", False) or bill_serial == 0 or bill_serial is None
        
        if not is_serial_na and bill_serial > 0:
            # Has valid serial number - use it directly (e.g., 7, 42, 156)
            return str(int(bill_serial))
        else:
            # Serial is NA/blank - find nearest property based on GPS and prefix with N
            nearest_serial = 0
            if valid_serials_with_gps and bill.get("latitude") and bill.get("longitude"):
                min_distance = float('inf')
                bill_lat = float(bill["latitude"])
                bill_lng = float(bill["longitude"])
                
                for vs in valid_serials_with_gps:
                    dist = ((vs["lat"] - bill_lat) ** 2 + (vs["lng"] - bill_lng) ** 2) ** 0.5
                    if dist < min_distance:
                        min_distance = dist
                        nearest_serial = vs["serial"]
            elif valid_serials_with_gps:
                nearest_serial = valid_serials_with_gps[0]["serial"]
            
            if nearest_serial > 0:
                return f"N{nearest_serial}"
            else:
                return "N/A"
    
    for emp_idx in range(employee_count):
        start_idx = emp_idx * bills_per_employee
        end_idx = min(start_idx + bills_per_employee, total_bills)
        
        if start_idx >= total_bills:
            break
        
        employee_bills = bills[start_idx:end_idx]
        
        output_filename = f"employee_{emp_idx + 1}_{colony or 'all'}_{timestamp}.pdf"
        output_path = UPLOAD_DIR / output_filename
        
        output_pdf = fitz.open()
        
        for bill in employee_bills:
            page_num = bill.get("page_number", 1) - 1
            if page_num < 0 or page_num >= len(src_pdf):
                continue
            
            output_pdf.insert_pdf(src_pdf, from_page=page_num, to_page=page_num)
            new_page = output_pdf[-1]
            
            # Get page rotation and dimensions
            rotation = new_page.rotation
            rect = new_page.rect
            
            # Get the serial number text
            sn_text = get_display_serial(bill)
            
            # Add Hindi message FIRST (left side), then serial number (right side)
            # Both at top with 50px padding
            is_self_certified = bill.get("self_certified", False)
            
            # Load font for Hindi + English support
            # Try Gargi font for proper Hindi rendering
            gargi_font = '/usr/share/fonts/truetype/Gargi/Gargi.ttf'
            samyak_font = '/usr/share/fonts/truetype/samyak-fonts/Samyak-Devanagari.ttf'
            freesans_font = '/usr/share/fonts/truetype/freefont/FreeSans.ttf'
            
            if os.path.exists(gargi_font):
                new_page.insert_font(fontname='gargi', fontbuffer=open(gargi_font, 'rb').read())
                font_name = 'gargi'
            elif os.path.exists(samyak_font):
                new_page.insert_font(fontname='samyak', fontbuffer=open(samyak_font, 'rb').read())
                font_name = 'samyak'
            elif os.path.exists(freesans_font):
                new_page.insert_font(fontname='freesans', fontbuffer=open(freesans_font, 'rb').read())
                font_name = 'freesans'
            else:
                font_name = 'helv'
            
            # Add serial number (RIGHT side)
            if rotation == 90:
                visual_point = fitz.Point(rect.width - 80, 50)
                internal_point = visual_point * new_page.derotation_matrix
                text_rotate = 90
            elif rotation == 270:
                visual_point = fitz.Point(80, rect.height - 50)
                internal_point = visual_point * new_page.derotation_matrix
                text_rotate = 270
            else:
                internal_point = fitz.Point(rect.width - 80, 50)
                text_rotate = 0
            
            new_page.insert_text(
                internal_point, 
                sn_text, 
                fontsize=sn_font_size, 
                color=sn_rgb, 
                fontname="helv",
                rotate=text_rotate
            )
        
        output_pdf.save(
            str(output_path),
            garbage=4,  # Maximum garbage collection
            deflate=True,  # Compress streams
            deflate_images=True,  # Compress images
            deflate_fonts=True   # Compress fonts
        )
        output_pdf.close()
        
        generated_files.append({
            "employee_number": emp_idx + 1,
            "filename": output_filename,
            "download_url": f"/api/uploads/{output_filename}",
            "bill_range": f"SN {employee_bills[0]['serial_number']} - {employee_bills[-1]['serial_number']}",
            "total_bills": len(employee_bills)
        })
    
    src_pdf.close()
    
    return {
        "message": f"Generated {len(generated_files)} employee PDFs",
        "total_bills": total_bills,
        "bills_per_employee": bills_per_employee,
        "files": generated_files
    }

@api_router.get("/admin/bills/map-data")
async def get_bills_map_data(
    batch_id: Optional[str] = None,
    colony: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get bill data for map display"""
    if current_user["role"] not in ADMIN_VIEW_ROLES:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    query = {}
    if batch_id and batch_id.strip():
        query["batch_id"] = batch_id
    if colony and colony.strip():
        query["colony"] = {"$regex": colony, "$options": "i"}
    
    # Only get bills with GPS coordinates
    query["latitude"] = {"$ne": None}
    query["longitude"] = {"$ne": None}
    
    bills = await db.bills.find(query, {
        "_id": 0,
        "id": 1,
        "serial_number": 1,
        "property_id": 1,
        "owner_name": 1,
        "mobile": 1,
        "colony": 1,
        "latitude": 1,
        "longitude": 1,
        "total_outstanding": 1,
        "category": 1
    }).sort("serial_number", 1).to_list(None)
    
    return {
        "bills": bills,
        "total": len(bills)
    }

@api_router.delete("/admin/bills/batch/{batch_id}")
async def delete_bill_batch(batch_id: str, current_user: dict = Depends(get_current_user)):
    """Delete a bill batch and all its bills"""
    if current_user["role"] not in ADMIN_ROLES:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Delete bills
    result = await db.bills.delete_many({"batch_id": batch_id})
    
    # Delete batch
    await db.batches.delete_one({"id": batch_id})
    
    return {"message": f"Deleted batch and {result.deleted_count} bills"}

@api_router.post("/admin/bills/delete-all")
async def delete_all_bills(
    batch_id: str = Form(None),
    colony: str = Form(None),
    current_user: dict = Depends(get_current_user)
):
    """Delete all bills matching the given filters. If no filters, deletes ALL bills."""
    if current_user["role"] not in ADMIN_ROLES:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    query = {}
    if batch_id and batch_id.strip():
        query["batch_id"] = batch_id
    if colony and colony.strip():
        query["colony"] = {"$regex": colony, "$options": "i"}
    
    # Get count first
    count = await db.bills.count_documents(query)
    
    if count == 0:
        return {"message": "No bills found to delete", "deleted_count": 0}
    
    # Delete the bills
    result = await db.bills.delete_many(query)
    
    # Update batch record counts if batch_id specified
    if batch_id and batch_id.strip():
        remaining = await db.bills.count_documents({"batch_id": batch_id})
        await db.batches.update_one(
            {"id": batch_id},
            {"$set": {"total_records": remaining}}
        )
        # If no bills left, delete the batch
        if remaining == 0:
            await db.batches.delete_one({"id": batch_id})
    
    return {
        "message": f"Successfully deleted {result.deleted_count} bills",
        "deleted_count": result.deleted_count
    }

@api_router.post("/admin/bills/copy-to-properties")
async def copy_bills_to_properties(
    batch_id: str = Form(None),
    colony: str = Form(None),
    skip_duplicates: str = Form("false"),
    skip_vacant_plots: str = Form("false"),
    skip_duplicate_gps: str = Form("false"),
    current_user: dict = Depends(get_current_user)
):
    """Copy bill data to properties collection"""
    if current_user["role"] not in ADMIN_ROLES:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Parse options
    should_skip_duplicates = skip_duplicates.lower() == "true"
    should_skip_vacant = skip_vacant_plots.lower() == "true"
    should_skip_duplicate_gps = skip_duplicate_gps.lower() == "true"
    
    query = {}
    if batch_id and batch_id.strip():
        query["batch_id"] = batch_id
    if colony and colony.strip():
        query["colony"] = {"$regex": colony, "$options": "i"}
    
    # Get bills to copy
    bills = await db.bills.find(query, {"_id": 0}).sort("serial_number", 1).to_list(None)
    
    if not bills:
        raise HTTPException(status_code=404, detail="No bills found to copy")
    
    # Get existing property_ids to check for duplicates (only if skip_duplicates is true)
    existing_property_ids = set()
    existing_keys = set()
    
    if should_skip_duplicates:
        existing_properties = await db.properties.find({}, {"property_id": 1, "owner_name": 1, "mobile": 1, "_id": 0}).to_list(None)
        existing_property_ids = set(p.get("property_id", "") for p in existing_properties)
        for p in existing_properties:
            owner = (p.get("owner_name") or "").strip().upper()
            mobile = (p.get("mobile") or "").strip()
            if owner and mobile:
                existing_keys.add(f"{owner}_{mobile}")
    
    # Track GPS coordinates to skip duplicates
    seen_gps = set()
    skipped_duplicate_gps = 0
    
    # Create a new batch for properties
    prop_batch_id = str(uuid.uuid4())
    prop_batch_name = f"Bills Import {colony or 'All'} - {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    
    prop_batch_doc = {
        "id": prop_batch_id,
        "name": prop_batch_name,
        "uploaded_by": current_user["id"],
        "uploaded_at": datetime.now(timezone.utc).isoformat(),
        "status": "ACTIVE",
        "total_records": 0,
        "source": "PDF_BILLS"
    }
    
    # First, get all valid serial numbers with GPS from bills for nearest-serial lookup
    valid_serials_with_gps = []
    for i, b in enumerate(bills):
        if not b.get("serial_na", False) and b.get("serial_number", 0) > 0 and b.get("latitude") and b.get("longitude"):
            valid_serials_with_gps.append({
                "serial": b["serial_number"],
                "lat": b["latitude"],
                "lng": b["longitude"]
            })
    
    # Vacant plot keywords - SAME as Generate PDF function
    def should_skip_for_property(bill):
        """Skip vacant plots and bills with no valid owner name - SAME as PDF generation"""
        owner = (bill.get("owner_name") or "").strip().lower()
        category = (bill.get("category") or "").strip().lower()
        
        # Skip if no owner name or invalid owner
        if not owner or owner in ['na', 'n/a', 'n.a.', '-', '--', 'nil', 'none']:
            return True
        
        # Skip vacant plots
        if "vacant" in category or "empty" in category:
            return True
        if "vacant" in owner or "empty plot" in owner or "खाली" in owner:
            return True
        
        return False
    
    # Load self-certified PIDs from database for matching
    self_certified_docs = await db.self_certified_pids.find({}, {"pid": 1, "_id": 0}).to_list(None)
    self_certified_pids = set(doc["pid"].upper() for doc in self_certified_docs)
    self_certified_count = 0
    not_self_certified_count = 0
    
    # Convert bills to properties
    properties = []
    skipped_duplicates = 0
    skipped_vacant = 0
    
    for i, bill in enumerate(bills):
        bill_prop_id = bill.get("property_id", "")
        
        # Skip vacant plots and invalid owner names (SAME logic as Generate PDF)
        if should_skip_vacant and should_skip_for_property(bill):
            skipped_vacant += 1
            continue
        
        # Skip duplicate GPS coordinates if option is enabled
        if should_skip_duplicate_gps:
            lat = bill.get("latitude")
            lng = bill.get("longitude")
            if lat and lng:
                # Round to 6 decimal places for comparison (about 0.1m precision)
                gps_key = f"{round(lat, 6)}_{round(lng, 6)}"
                if gps_key in seen_gps:
                    skipped_duplicate_gps += 1
                    continue
                seen_gps.add(gps_key)
        
        # Only check duplicates if option is enabled
        if should_skip_duplicates:
            # Check for duplicate by property_id
            if bill_prop_id and bill_prop_id in existing_property_ids:
                skipped_duplicates += 1
                continue
            
            # Check for duplicate by owner_name + mobile
            owner = (bill.get("owner_name") or "").strip().upper()
            mobile = (bill.get("mobile") or "").strip()
            if owner and mobile:
                key = f"{owner}_{mobile}"
                if key in existing_keys:
                    skipped_duplicates += 1
                    continue
                existing_keys.add(key)
        
        # Use the actual BillSrNo from PDF, or mark as N/A
        bill_serial = bill.get("serial_number", 0)
        is_serial_na = bill.get("serial_na", False) or bill_serial == 0
        
        # Format N/A serials as N-X where X is the nearest valid serial BY GPS
        if is_serial_na:
            nearest_serial = 0
            if valid_serials_with_gps and bill.get("latitude") and bill.get("longitude"):
                min_distance = float('inf')
                bill_lat = bill["latitude"]
                bill_lng = bill["longitude"]
                
                for vs in valid_serials_with_gps:
                    dist = ((vs["lat"] - bill_lat) ** 2 + (vs["lng"] - bill_lng) ** 2) ** 0.5
                    if dist < min_distance:
                        min_distance = dist
                        nearest_serial = vs["serial"]
            elif valid_serials_with_gps:
                nearest_serial = valid_serials_with_gps[0]["serial"]
            
            bill_sr_no_display = f"N{nearest_serial}"
        else:
            bill_sr_no_display = str(bill_serial)
        
        # Check if this property is self-certified
        is_self_certified = bill_prop_id.upper() in self_certified_pids if bill_prop_id else False
        if is_self_certified:
            self_certified_count += 1
        else:
            not_self_certified_count += 1
        
        prop = {
            "id": str(uuid.uuid4()),
            "batch_id": prop_batch_id,
            "serial_number": bill_serial if not is_serial_na else 0,
            "serial_na": is_serial_na,
            "bill_sr_no": bill_sr_no_display,
            "property_id": bill_prop_id if bill_prop_id else str(uuid.uuid4())[:8].upper(),
            "old_property_id": bill.get("old_property_id", ""),
            "owner_name": bill.get("owner_name", "Unknown"),
            "mobile": bill.get("mobile", ""),
            "address": bill.get("plot_address", ""),
            "colony": bill.get("colony", ""),
            "ward": bill.get("colony", ""),  # Use colony as ward
            "latitude": bill.get("latitude"),
            "longitude": bill.get("longitude"),
            "total_area": bill.get("total_area", ""),
            "category": bill.get("category", ""),
            "amount": bill.get("total_outstanding", "0"),
            "financial_year": bill.get("financial_year", "2025-2026"),
            "assigned_employee_id": None,
            "assigned_employee_name": None,
            "status": "Pending",
            "self_certified": is_self_certified,  # NEW: Self-certification status
            "created_at": datetime.now(timezone.utc).isoformat(),
            "source_bill_id": bill.get("id")  # Reference to original bill
        }
        properties.append(prop)
        
        # Add to existing set to prevent duplicates within same batch
        if bill_prop_id:
            existing_property_ids.add(bill_prop_id)
    
    # Build detailed message
    msg_parts = [f"Successfully added {len(properties)} bills to properties"]
    if self_certified_count > 0:
        msg_parts.append(f"{self_certified_count} self-certified")
    if not_self_certified_count > 0:
        msg_parts.append(f"{not_self_certified_count} not self-certified")
    if skipped_duplicates > 0:
        msg_parts.append(f"Skipped {skipped_duplicates} duplicates")
    if skipped_vacant > 0:
        msg_parts.append(f"Skipped {skipped_vacant} vacant plots")
    if skipped_duplicate_gps > 0:
        msg_parts.append(f"Skipped {skipped_duplicate_gps} duplicate GPS")
    
    # Save detailed stats to batch
    add_to_properties_stats = {
        "total_added": len(properties),
        "self_certified": self_certified_count,
        "not_self_certified": not_self_certified_count,
        "skipped_duplicates": skipped_duplicates,
        "skipped_vacant": skipped_vacant,
        "skipped_duplicate_gps": skipped_duplicate_gps,
        "message": ". ".join(msg_parts) + "."
    }
    
    # Insert properties
    if properties:
        await db.properties.insert_many(properties)
        prop_batch_doc["total_records"] = len(properties)
        prop_batch_doc["add_to_properties_stats"] = add_to_properties_stats
        await db.batches.update_one({"id": prop_batch_id}, {"$set": {"total_records": len(properties)}})
    
    # Also update the source PDF batch with these stats
    if batch_id:
        await db.batches.update_one(
            {"id": batch_id},
            {"$set": {"add_to_properties_stats": add_to_properties_stats}}
        )
    
    await db.batches.insert_one(prop_batch_doc)
    
    return {
        "message": ". ".join(msg_parts) + ".",
        "batch_id": prop_batch_id,
        "batch_name": prop_batch_name,
        "total_added": len(properties),
        "self_certified": self_certified_count,
        "not_self_certified": not_self_certified_count,
        "skipped_duplicates": skipped_duplicates,
        "skipped_vacant": skipped_vacant,
        "skipped_duplicate_gps": skipped_duplicate_gps
    }

@api_router.post("/admin/bills/split-by-employees")
async def split_bills_by_specific_employees(
    batch_id: str = Form(None),
    colony: str = Form(None),
    employee_ids: str = Form(...),  # Comma-separated employee IDs
    sn_font_size: int = Form(48),
    sn_color: str = Form("red"),
    current_user: dict = Depends(get_current_user)
):
    """Split bills among specific employees and generate separate PDFs"""
    if current_user["role"] not in ADMIN_ROLES:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Parse employee IDs
    emp_ids = [e.strip() for e in employee_ids.split(",") if e.strip()]
    
    if not emp_ids:
        raise HTTPException(status_code=400, detail="At least one employee must be selected")
    
    # Verify employees exist
    employees = []
    for emp_id in emp_ids:
        emp = await db.users.find_one({"id": emp_id}, {"_id": 0, "id": 1, "name": 1, "username": 1})
        if emp:
            employees.append(emp)
    
    if not employees:
        raise HTTPException(status_code=404, detail="No valid employees found")
    
    query = {}
    if batch_id and batch_id.strip():
        query["batch_id"] = batch_id
    if colony and colony.strip():
        query["colony"] = {"$regex": colony, "$options": "i"}
    
    # Get arranged bills
    bills = await db.bills.find(query, {"_id": 0}).sort("serial_number", 1).to_list(None)
    
    if not bills:
        raise HTTPException(status_code=404, detail="No bills found")
    
    # Get original PDF
    batch = await db.batches.find_one({"id": bills[0]["batch_id"]})
    if not batch or not batch.get("pdf_filename"):
        raise HTTPException(status_code=404, detail="Original PDF not found")
    
    original_pdf_path = UPLOAD_DIR / batch["pdf_filename"]
    if not original_pdf_path.exists():
        raise HTTPException(status_code=404, detail="Original PDF file not found")
    
    # Calculate bills per employee
    total_bills = len(bills)
    employee_count = len(employees)
    bills_per_employee = math.ceil(total_bills / employee_count)
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    generated_files = []
    
    # Color mapping
    color_map = {
        "red": (1, 0, 0),
        "blue": (0, 0, 1),
        "green": (0, 0.5, 0),
        "black": (0, 0, 0),
        "orange": (1, 0.5, 0)
    }
    sn_rgb = color_map.get(sn_color.lower(), (1, 0, 0))
    
    src_pdf = fitz.open(str(original_pdf_path))
    
    # Build serial number lookup from ALL bills for N/A serials
    valid_serials_with_gps = []
    for b in bills:
        if not b.get("serial_na", False) and b.get("serial_number", 0) > 0 and b.get("latitude") and b.get("longitude"):
            valid_serials_with_gps.append({
                "serial": b["serial_number"],
                "lat": b["latitude"],
                "lng": b["longitude"]
            })
    
    def get_display_serial(bill):
        """Get display serial number:
        - If bill has a valid serial_number (not 0, not NA) → use that number (e.g., 7, 42)
        - If serial is NA/blank/0 → find nearest property with valid serial based on GPS and prefix with N (e.g., N7)
        """
        bill_serial = bill.get("serial_number") or 0
        is_serial_na = bill.get("serial_na", False) or bill_serial == 0 or bill_serial is None
        
        if not is_serial_na and bill_serial > 0:
            # Has valid serial number - use it directly (e.g., 7, 42, 156)
            return str(int(bill_serial))
        else:
            # Serial is NA/blank - find nearest property based on GPS and prefix with N
            nearest_serial = 0
            if valid_serials_with_gps and bill.get("latitude") and bill.get("longitude"):
                min_distance = float('inf')
                bill_lat = float(bill["latitude"])
                bill_lng = float(bill["longitude"])
                
                for vs in valid_serials_with_gps:
                    dist = ((vs["lat"] - bill_lat) ** 2 + (vs["lng"] - bill_lng) ** 2) ** 0.5
                    if dist < min_distance:
                        min_distance = dist
                        nearest_serial = vs["serial"]
            elif valid_serials_with_gps:
                nearest_serial = valid_serials_with_gps[0]["serial"]
            
            if nearest_serial > 0:
                return f"N{nearest_serial}"
            else:
                return "N/A"
    
    for emp_idx, emp in enumerate(employees):
        start_idx = emp_idx * bills_per_employee
        end_idx = min(start_idx + bills_per_employee, total_bills)
        
        if start_idx >= total_bills:
            break
        
        employee_bills = bills[start_idx:end_idx]
        
        # Use employee name in filename (sanitize for filename)
        emp_name_safe = re.sub(r'[^\w\-_]', '_', emp.get('name', f'emp_{emp_idx+1}'))
        output_filename = f"{emp_name_safe}_{colony or 'all'}_{timestamp}.pdf"
        output_path = UPLOAD_DIR / output_filename
        
        output_pdf = fitz.open()
        
        for bill in employee_bills:
            page_num = bill.get("page_number", 1) - 1
            if page_num < 0 or page_num >= len(src_pdf):
                continue
            
            # Simply copy the page as-is
            output_pdf.insert_pdf(src_pdf, from_page=page_num, to_page=page_num)
            new_page = output_pdf[-1]
            
            # Get page rotation and dimensions
            rotation = new_page.rotation
            rect = new_page.rect
            
            # Get the serial number text
            sn_text = get_display_serial(bill)
            
            # Add Hindi message FIRST (left side), then serial number (right side)
            # Both at top with 50px padding
            is_self_certified = bill.get("self_certified", False)
            
            # Load font for Hindi + English support
            # Try Gargi font for proper Hindi rendering
            gargi_font = '/usr/share/fonts/truetype/Gargi/Gargi.ttf'
            samyak_font = '/usr/share/fonts/truetype/samyak-fonts/Samyak-Devanagari.ttf'
            freesans_font = '/usr/share/fonts/truetype/freefont/FreeSans.ttf'
            
            if os.path.exists(gargi_font):
                new_page.insert_font(fontname='gargi', fontbuffer=open(gargi_font, 'rb').read())
                font_name = 'gargi'
            elif os.path.exists(samyak_font):
                new_page.insert_font(fontname='samyak', fontbuffer=open(samyak_font, 'rb').read())
                font_name = 'samyak'
            elif os.path.exists(freesans_font):
                new_page.insert_font(fontname='freesans', fontbuffer=open(freesans_font, 'rb').read())
                font_name = 'freesans'
            else:
                font_name = 'helv'
            
            # Add serial number (RIGHT side)
            if rotation == 90:
                visual_point = fitz.Point(rect.width - 80, 50)
                internal_point = visual_point * new_page.derotation_matrix
                text_rotate = 90
            elif rotation == 270:
                visual_point = fitz.Point(80, rect.height - 50)
                internal_point = visual_point * new_page.derotation_matrix
                text_rotate = 270
            else:
                internal_point = fitz.Point(rect.width - 80, 50)
                text_rotate = 0
            
            new_page.insert_text(
                internal_point, 
                sn_text, 
                fontsize=sn_font_size, 
                color=sn_rgb, 
                fontname="helv",
                rotate=text_rotate
            )
            
            # Add note for non-self-certified properties - using pre-generated image
            if not is_self_certified:
                # Use pre-generated Hindi note image
                note_img_path = "/tmp/hindi_note_cached.png"
                
                # Generate image only if it doesn't exist
                if not os.path.exists(note_img_path):
                    import subprocess
                    import tempfile
                    
                    hindi_note_html = '''<!DOCTYPE html>
<html><head><meta charset="UTF-8">
<style>body{margin:0;padding:0;font-family:'Noto Sans Devanagari','Lohit Devanagari',sans-serif;font-size:34px;color:#cc0000;background:transparent;white-space:nowrap;}</style>
</head><body>Note : आप अपनी Property ID को सेल्फ सर्टिफाइड करवाए, जिससे कि आपकी Property के साथ कोई छेड़ -छाड़ ना कर सके।</body></html>'''
                    
                    with tempfile.NamedTemporaryFile(mode='w', suffix='.html', delete=False, encoding='utf-8') as f:
                        f.write(hindi_note_html)
                        html_path = f.name
                    
                    try:
                        subprocess.run([
                            'xvfb-run', '--auto-servernum', 'wkhtmltoimage',
                            '--encoding', 'utf-8', '--width', '1300', '--height', '70', '--quality', '100',
                            html_path, note_img_path
                        ], capture_output=True, timeout=30)
                    finally:
                        if os.path.exists(html_path):
                            os.unlink(html_path)
                
                if os.path.exists(note_img_path):
                    try:
                        # Insert image into PDF - left aligned, 300px from bottom
                        if rotation == 90:
                            # Y=5 for left align, 300px padding from bottom
                            img_rect = fitz.Rect(477, 5, 542, 590)
                        elif rotation == 270:
                            img_rect = fitz.Rect(rect.width - 542, rect.height - 590, rect.width - 477, rect.height - 5)
                        else:
                            img_rect = fitz.Rect(5, rect.height - 390, rect.width - 5, rect.height - 325)
                        
                        new_page.insert_image(img_rect, filename=note_img_path, rotate=rotation)
                    except Exception as e:
                        logger.error(f"Error inserting Hindi note image: {e}")
        
        output_pdf.save(
            str(output_path),
            garbage=4,  # Maximum garbage collection
            deflate=True,  # Compress streams
            deflate_images=True,  # Compress images
            deflate_fonts=True   # Compress fonts
        )
        output_pdf.close()
        
        generated_files.append({
            "employee_id": emp["id"],
            "employee_name": emp.get("name", emp.get("username", f"Employee {emp_idx+1}")),
            "filename": output_filename,
            "download_url": f"/api/uploads/{output_filename}",
            "bill_range": f"SR {employee_bills[0]['serial_number']} - {employee_bills[-1]['serial_number']}",
            "total_bills": len(employee_bills)
        })
    
    src_pdf.close()
    
    return {
        "message": f"Generated PDFs for {len(generated_files)} employees",
        "total_bills": total_bills,
        "bills_per_employee": bills_per_employee,
        "files": generated_files
    }

# ============== GENERATED PDFs MANAGEMENT ==============

class GeneratedPdfRecord(BaseModel):
    colony: str
    filename: str
    download_url: str
    pdf_type: str  # "arranged_bills", "survey_report", "property_list"
    total_records: int
    file_size: Optional[int] = None

@api_router.post("/admin/generated-pdfs/save")
async def save_generated_pdf(
    colony: str = Form(...),
    filename: str = Form(...),
    download_url: str = Form(...),
    pdf_type: str = Form("arranged_bills"),
    total_records: int = Form(0),
    current_user: dict = Depends(get_current_user)
):
    """Save a generated PDF record to database for later download"""
    if current_user["role"] not in ADMIN_ROLES:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Get file size if exists
    file_path = UPLOAD_DIR / filename
    file_size = file_path.stat().st_size if file_path.exists() else 0
    
    pdf_doc = {
        "id": str(uuid.uuid4()),
        "colony": colony,
        "filename": filename,
        "download_url": download_url,
        "pdf_type": pdf_type,
        "total_records": total_records,
        "file_size": file_size,
        "created_by": current_user["id"],
        "created_by_name": current_user["name"],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.generated_pdfs.insert_one(pdf_doc)
    
    return {
        "message": "PDF record saved",
        "id": pdf_doc["id"],
        "filename": filename
    }

@api_router.get("/admin/generated-pdfs")
async def list_generated_pdfs(
    colony: Optional[str] = None,
    pdf_type: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """List all generated PDFs, optionally filtered by colony"""
    if current_user["role"] not in ADMIN_VIEW_ROLES:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    query = {}
    if colony and colony.strip():
        query["colony"] = {"$regex": f"^{colony}$", "$options": "i"}
    if pdf_type and pdf_type.strip():
        query["pdf_type"] = pdf_type
    
    pdfs = await db.generated_pdfs.find(query, {"_id": 0}).sort("created_at", -1).to_list(100)
    
    # Check which files still exist
    for pdf in pdfs:
        file_path = UPLOAD_DIR / pdf["filename"]
        pdf["file_exists"] = file_path.exists()
        if pdf["file_exists"] and not pdf.get("file_size"):
            pdf["file_size"] = file_path.stat().st_size
    
    return {"pdfs": pdfs, "total": len(pdfs)}

@api_router.get("/admin/generated-pdfs/by-colony")
async def get_pdfs_by_colony(current_user: dict = Depends(get_current_user)):
    """Get grouped list of generated PDFs by colony"""
    if current_user["role"] not in ADMIN_VIEW_ROLES:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    pipeline = [
        {"$match": {}},
        {"$sort": {"created_at": -1}},
        {"$group": {
            "_id": "$colony",
            "latest_pdf": {"$first": "$$ROOT"},
            "total_pdfs": {"$sum": 1}
        }},
        {"$sort": {"_id": 1}}
    ]
    
    result = await db.generated_pdfs.aggregate(pipeline).to_list(None)
    
    colonies_with_pdfs = []
    for r in result:
        pdf = r["latest_pdf"]
        file_path = UPLOAD_DIR / pdf["filename"]
        colonies_with_pdfs.append({
            "colony": r["_id"],
            "total_pdfs": r["total_pdfs"],
            "latest_filename": pdf["filename"],
            "latest_download_url": pdf["download_url"],
            "latest_created_at": pdf["created_at"],
            "latest_total_records": pdf.get("total_records", 0),
            "file_exists": file_path.exists()
        })
    
    return {"colonies": colonies_with_pdfs}

@api_router.delete("/admin/generated-pdfs/{pdf_id}")
async def delete_generated_pdf(pdf_id: str, current_user: dict = Depends(get_current_user)):
    """Delete a generated PDF record and optionally the file"""
    if current_user["role"] not in ADMIN_ROLES:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    pdf = await db.generated_pdfs.find_one({"id": pdf_id}, {"_id": 0})
    if not pdf:
        raise HTTPException(status_code=404, detail="PDF record not found")
    
    # Delete file if exists
    file_path = UPLOAD_DIR / pdf["filename"]
    if file_path.exists():
        file_path.unlink()
    
    await db.generated_pdfs.delete_one({"id": pdf_id})
    
    return {"message": "PDF record deleted", "filename": pdf["filename"]}

@api_router.get("/admin/generated-pdfs/download/{filename}")
async def download_generated_pdf(filename: str, current_user: dict = Depends(get_current_user)):
    """Download a previously generated PDF file"""
    if current_user["role"] not in ADMIN_VIEW_ROLES:
        raise HTTPException(status_code=403, detail="Access denied")
    
    file_path = UPLOAD_DIR / filename
    if not file_path.exists():
        raise HTTPException(status_code=404, detail=f"File not found: {filename}")
    
    return FileResponse(
        path=str(file_path),
        media_type='application/pdf',
        filename=filename,
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"'
        }
    )

# ============== INITIALIZATION ==============

@api_router.get("/")
async def root():
    return {"message": "NSTU Property Tax Manager API"}

# ==============================================
# SELF CERTIFICATION ENDPOINTS
# ==============================================

@api_router.post("/admin/upload-self-certification")
async def upload_self_certification(
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    """Upload self-certification Excel/CSV file to store certified property IDs"""
    if current_user["role"] not in ADMIN_ROLES:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    if not file.filename.endswith(('.xlsx', '.xls', '.csv')):
        raise HTTPException(status_code=400, detail="Please upload an Excel (.xlsx, .xls) or CSV file")
    
    try:
        # Read the file
        contents = await file.read()
        
        if file.filename.endswith('.csv'):
            df = pd.read_csv(io.BytesIO(contents))
        else:
            df = pd.read_excel(io.BytesIO(contents))
        
        # Find the PID column (could be 'PID (C)' or similar)
        pid_column = None
        for col in df.columns:
            if 'PID' in col.upper() or 'PROPERTY' in col.upper() and 'ID' in col.upper():
                pid_column = col
                break
        
        if not pid_column:
            # Try first column if no PID column found
            pid_column = df.columns[0] if len(df.columns) > 0 else None
        
        if not pid_column:
            raise HTTPException(status_code=400, detail="Could not find Property ID column in Excel file")
        
        # Extract unique PIDs
        pids = df[pid_column].dropna().astype(str).str.strip().str.upper().unique().tolist()
        
        if not pids:
            raise HTTPException(status_code=400, detail="No property IDs found in the file")
        
        # Store in database - create or update the self_certified_pids collection
        # First, get existing PIDs to avoid duplicates
        existing = await db.self_certified_pids.find({}, {"pid": 1, "_id": 0}).to_list(None)
        existing_pids = set(p["pid"] for p in existing)
        
        # Only insert new PIDs
        new_pids = [pid for pid in pids if pid not in existing_pids]
        
        if new_pids:
            docs = [{"pid": pid, "uploaded_at": datetime.now(timezone.utc).isoformat()} for pid in new_pids]
            await db.self_certified_pids.insert_many(docs)
        
        # Create index for fast lookups
        await db.self_certified_pids.create_index("pid")
        
        return {
            "message": f"Uploaded {len(new_pids)} new self-certified PIDs. {len(existing_pids)} already existed.",
            "total_in_file": len(pids),
            "new_added": len(new_pids),
            "already_existed": len(pids) - len(new_pids),
            "total_in_database": len(existing_pids) + len(new_pids)
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error processing file: {str(e)}")

@api_router.get("/admin/self-certification-stats")
async def get_self_certification_stats(current_user: dict = Depends(get_current_user)):
    """Get statistics about self-certified properties"""
    if current_user["role"] not in ADMIN_ROLES:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    total_certified = await db.self_certified_pids.count_documents({})
    
    return {
        "total_self_certified_pids": total_certified
    }

@api_router.delete("/admin/clear-self-certification")
async def clear_self_certification(current_user: dict = Depends(get_current_user)):
    """Clear all self-certification data (use with caution)"""
    if current_user["role"] != "ADMIN":
        raise HTTPException(status_code=403, detail="Super Admin access required")
    
    result = await db.self_certified_pids.delete_many({})
    
    return {
        "message": f"Cleared {result.deleted_count} self-certified PIDs"
    }

@api_router.post("/init-admin")
async def init_admin():
    """Initialize default admin user if not exists"""
    existing = await db.users.find_one({"username": "admin"})
    if existing:
        return {"message": "Admin already exists"}
    
    admin_doc = {
        "id": str(uuid.uuid4()),
        "username": "admin",
        "password_hash": hash_password("admin123"),
        "name": "Super Admin",
        "role": "ADMIN",
        "assigned_area": None,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.users.insert_one(admin_doc)
    return {"message": "Admin user created", "username": "admin", "password": "admin123"}

# Include the router
app.include_router(api_router)

# CORS
app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

# Logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
